
import pandas as pd
def avg_ignore_zeros_cols(row, cols):
    """
    Average of columns in row ignoring zeros/NaN, and ignoring the earliest week column.
    """
    use_cols = _week_cols_excluding_first(row.to_frame().T, cols)
    vals = []
    for c in use_cols:
        v = row.get(c, np.nan)
        if pd.isna(v):
            continue
        try:
            fv = float(v)
        except Exception:
            continue
        if fv == 0:
            continue
        vals.append(fv)
    return float(np.mean(vals)) if vals else 0.0


def _week_cols_excluding_first(df, week_cols):
    """
    Remove the earliest week column from week_cols (to ignore partial first week).
    Uses parsed week start date from the column name when possible.
    """
    if not week_cols:
        return week_cols
    parsed = [pd.to_datetime(c, errors="coerce") for c in week_cols]
    if all(pd.isna(p) for p in parsed):
        return week_cols[1:] if len(week_cols) > 1 else week_cols
    pairs = [(c, p) for c, p in zip(week_cols, parsed) if pd.notna(p)]
    if not pairs:
        return week_cols[1:] if len(week_cols) > 1 else week_cols
    earliest = min(pairs, key=lambda x: x[1])[0]
    return [c for c in week_cols if c != earliest]


import re
from pathlib import Path
from datetime import date, timedelta

import numpy as np
import pandas as pd
import streamlit as st

MONTH_NAME_TO_NUM = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12
}

AVG_WINDOW_OPTIONS = ["4 weeks","5 weeks","6 weeks","7 weeks","8 weeks","9 weeks","10 weeks","11 weeks","12 weeks",
                      "January","February","March","April","May","June","July","August","September","October","November","December"]

def resolve_avg_use(avg_window, use_cols, current_year):
    """Return which week columns to use for averaging. Month choices are within current_year only.

    Notes:
      - use_cols may be a list of datetime.date objects or strings.
      - Convert to a Series so we can safely use .dt accessors.
    """
    if not use_cols:
        return []
    if isinstance(avg_window, str) and avg_window in MONTH_NAME_TO_NUM:
        mnum = MONTH_NAME_TO_NUM[avg_window]
        dates = pd.to_datetime(pd.Series(list(use_cols)), errors="coerce")
        mask = (dates.dt.year == int(current_year)) & (dates.dt.month == int(mnum))
        cols = [c for c, ok in zip(use_cols, mask.fillna(False).tolist()) if ok]
        return cols
    # rolling weeks like '8 weeks'
    if isinstance(avg_window, str) and "week" in avg_window:
        try:
            n = int(avg_window.split()[0])
        except Exception:
            n = 4
        return use_cols[-n:] if len(use_cols) >= n else use_cols
    return use_cols


APP_TITLE = "Sales Dashboard (Vendor Map + Weekly Sheets)"
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

DEFAULT_VENDOR_MAP = DATA_DIR / "vendor_map.xlsx"
DEFAULT_SALES_STORE = DATA_DIR / "sales_store.csv"
DEFAULT_PRICE_HISTORY = DATA_DIR / "price_history.csv"


# Year locks (prevent accidental edits to closed years)
DEFAULT_YEAR_LOCKS = DATA_DIR / "year_locks.json"

def load_year_locks() -> set[int]:
    try:
        if DEFAULT_YEAR_LOCKS.exists():
            obj = json.loads(DEFAULT_YEAR_LOCKS.read_text(encoding="utf-8"))
            years = obj.get("locked_years", [])
            return set(int(y) for y in years)
    except Exception:
        pass
    return set()

def save_year_locks(locked_years: set[int]) -> None:
    try:
        DEFAULT_YEAR_LOCKS.write_text(json.dumps({"locked_years": sorted(list(locked_years))}, indent=2), encoding="utf-8")
    except Exception:
        return

def overwrite_sales_rows(target_year: int, retailers: set[str]) -> None:
    """Remove rows from sales_store.csv for the given year + retailers."""
    if not DEFAULT_SALES_STORE.exists():
        return
    try:
        cur = pd.read_csv(DEFAULT_SALES_STORE)
        cur["StartDate"] = pd.to_datetime(cur.get("StartDate"), errors="coerce")
        cur["Retailer"] = cur.get("Retailer", "").map(_normalize_retailer)
        retailers_n = {_normalize_retailer(r) for r in retailers}
        keep = ~((cur["StartDate"].dt.year == int(target_year)) & (cur["Retailer"].isin(retailers_n)))
        cur2 = cur[keep].copy()
        cur2.to_csv(DEFAULT_SALES_STORE, index=False)
    except Exception:
        return

# -------------------------
# Normalization
# -------------------------
def _normalize_retailer(x: str) -> str:
    if x is None:
        return ""
    x = str(x).strip()
    aliases = {
        "home depot": "Depot",
        "depot": "Depot",
        "the home depot": "Depot",
        "lowes": "Lowe's",
        "lowe's": "Lowe's",
        "tractor supply": "Tractor Supply",
        "tsc": "Tractor Supply",
        "amazon": "Amazon",
    }
    key = re.sub(r"\s+", " ", x.lower()).strip()
    return aliases.get(key, x)

def _normalize_sku(x: str) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

# -------------------------
# Formatting
# -------------------------
def fmt_currency(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    s = f"${abs(v):,.2f}"
    return f"({s})" if v < 0 else s

def fmt_int(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    return f"{int(round(v)):,.0f}"

def fmt_2(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    return f"{v:,.2f}"

def _color(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "inherit"
    try:
        v = float(v)
    except Exception:
        return "inherit"
    if v > 0:
        return "green"
    if v < 0:
        return "red"
    return "inherit"

def _table_height(df: pd.DataFrame, row_px: int = 32, header_px: int = 38, max_px: int = 1100) -> int:
    if df is None:
        return 220
    n = int(df.shape[0])
    h = header_px + (n + 1) * row_px
    return int(min(max(h, 220), max_px))

def style_currency_cols(df: pd.DataFrame, diff_cols=None):
    diff_cols = diff_cols or []
    sty = df.style
    # format all non-first columns as currency
    first = df.columns[0]
    fmt = {c: (lambda v: fmt_currency(v)) for c in df.columns if c != first}
    sty = sty.format(fmt)
    for c in diff_cols:
        if c in df.columns:
            sty = sty.applymap(lambda v: f"color: {_color(v)};", subset=[c])
    return sty

# -------------------------
# Vendor map
# -------------------------
def load_vendor_map(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    cols = {c.lower().strip(): c for c in df.columns}

    def pick(name, fallbacks):
        for k in [name] + fallbacks:
            if k in cols:
                return cols[k]
        return None

    c_retail = pick("retailer", [])
    c_sku = pick("sku", ["item", "item sku"])
    c_vendor = pick("vendor", ["supplier"])
    c_price = pick("price", ["unit price", "cost"])

    out = pd.DataFrame({
        "Retailer": df[c_retail] if c_retail else "",
        "SKU": df[c_sku] if c_sku else "",
        "Vendor": df[c_vendor] if c_vendor else "",
        "Price": df[c_price] if c_price else np.nan,
    })

    out["Retailer"] = out["Retailer"].map(_normalize_retailer)
    out["SKU"] = out["SKU"].map(_normalize_sku)
    out["Vendor"] = out["Vendor"].astype(str).str.strip()
    out["Price"] = pd.to_numeric(out["Price"], errors="coerce")

    # preserve order per retailer
    out["MapOrder"] = 0
    for r, grp in out.groupby("Retailer", sort=False):
        for j, ix in enumerate(grp.index.tolist()):
            out.loc[ix, "MapOrder"] = j

    return out

# -------------------------
# Sales store
# -------------------------
def load_sales_store() -> pd.DataFrame:
    if DEFAULT_SALES_STORE.exists():
        df = pd.read_csv(DEFAULT_SALES_STORE)
        for c in ["StartDate", "EndDate"]:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c], errors="coerce")
        df["Retailer"] = df["Retailer"].map(_normalize_retailer)
        df["SKU"] = df["SKU"].map(_normalize_sku)
        df["Units"] = pd.to_numeric(df["Units"], errors="coerce").fillna(0.0)
        if "UnitPrice" in df.columns:
            df["UnitPrice"] = pd.to_numeric(df["UnitPrice"], errors="coerce")
        else:
            df["UnitPrice"] = np.nan
        return df
    return pd.DataFrame(columns=["Retailer","SKU","Units","UnitPrice","StartDate","EndDate","SourceFile"])



# -------------------------
# Price history (effective dating)
# -------------------------
def _normalize_price_retailer(x):
    x = "" if x is None else str(x).strip()
    if x == "" or x.lower() in {"all","*", "any"}:
        return "*"
    return _normalize_retailer(x)

def load_price_history() -> pd.DataFrame:
    """
    Returns columns: Retailer, SKU, Price, StartDate (datetime64)
    Retailer="*" means applies to all retailers for that SKU.
    """
    if DEFAULT_PRICE_HISTORY.exists():
        ph = pd.read_csv(DEFAULT_PRICE_HISTORY)
        # flexible column names
        colmap = {c.lower(): c for c in ph.columns}
        sku_col = colmap.get("sku") or colmap.get("sku#") or colmap.get("skunumber") or colmap.get("skuid")
        price_col = colmap.get("price") or colmap.get("unitprice") or colmap.get("unit_price")
        date_col = colmap.get("startdate") or colmap.get("start_date") or colmap.get("effective_date") or colmap.get("date")
        ret_col = colmap.get("retailer")

        if sku_col:
            ph["SKU"] = ph[sku_col].map(_normalize_sku)
        else:
            ph["SKU"] = ""
        if price_col:
            ph["Price"] = pd.to_numeric(ph[price_col], errors="coerce")
        else:
            ph["Price"] = np.nan
        if date_col:
            ph["StartDate"] = pd.to_datetime(ph[date_col], errors="coerce")
        else:
            ph["StartDate"] = pd.NaT
        if ret_col:
            ph["Retailer"] = ph[ret_col].map(_normalize_price_retailer)
        else:
            ph["Retailer"] = "*"

        ph = ph[["Retailer","SKU","Price","StartDate"]].dropna(subset=["SKU","Price","StartDate"])
        ph = ph.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
        return ph
    return pd.DataFrame(columns=["Retailer","SKU","Price","StartDate"])

def save_price_history(ph: pd.DataFrame) -> None:
    ph2 = ph.copy()
    ph2["StartDate"] = pd.to_datetime(ph2["StartDate"], errors="coerce")
    ph2 = ph2.dropna(subset=["Retailer","SKU","Price","StartDate"])
    ph2["Retailer"] = ph2["Retailer"].map(_normalize_price_retailer)
    ph2["SKU"] = ph2["SKU"].map(_normalize_sku)
    ph2["Price"] = pd.to_numeric(ph2["Price"], errors="coerce")
    ph2 = ph2.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
    ph2.to_csv(DEFAULT_PRICE_HISTORY, index=False)



def _prepare_price_history_upload(new_rows: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Normalize a price history upload. Returns:
      - normalized rows to consider (SKU, Retailer, StartDate, Price)
      - rows ignored (for reporting)

    Rules:
      - Price blank/NaN => ignored
      - Price <= 0 => ignored (treated as blank)
      - Missing SKU or StartDate => ignored

    Column name matching is forgiving (spaces/underscores/case).
    Accepts: SKU, Retailer, Price, StartDate (or "Start Date", "Effective Date", etc.)
    """
    n = new_rows.copy()

    def norm_key(s: str) -> str:
        s = str(s).strip().lower()
        # keep only alphanumerics to make "Start Date" == "start_date" == "StartDate"
        return re.sub(r"[^a-z0-9]+", "", s)

    cols = {norm_key(c): c for c in n.columns}

    def pick(*keys):
        for k in keys:
            if k in cols:
                return cols[k]
        return None

    sku_col = pick("sku", "sku#", "skunumber", "skuid", "itemsku")
    price_col = pick("price", "unitprice", "unit_price")
    date_col = pick("startdate", "start_date", "startdateeffective", "effectivedate", "effective_date", "start", "date", "startdate1", "startdate2", "startdate3", "startdate4", "startdate5", "startdate6", "startdate7", "startdate8", "startdate9", "startdate10", "startdate11", "startdate12", "startdate13", "startdate14", "startdate15", "startdate16", "startdate17", "startdate18", "startdate19", "startdate20", "startdate21", "startdate22", "startdate23", "startdate24", "startdate25", "startdate26", "startdate27", "startdate28", "startdate29", "startdate30", "startdate31", "startdate32", "startdate33", "startdate34", "startdate35", "startdate36", "startdate37", "startdate38", "startdate39", "startdate40", "startdate41", "startdate42", "startdate43", "startdate44", "startdate45", "startdate46", "startdate47", "startdate48", "startdate49", "startdate50", "startdate51", "startdate52", "startdate53", "startdate54", "startdate55", "startdate56", "startdate57", "startdate58", "startdate59", "startdate60")
    # Common: "start date"
    if date_col is None:
        date_col = pick("startdate", "startdate", "startdate")  # no-op, just for clarity
        date_col = cols.get("startdate") or cols.get("startdate")  # no-op

    # Explicitly support "Start Date" / "Effective Date"
    if date_col is None:
        date_col = pick("startdate", "startdate")  # still none
    if date_col is None:
        date_col = pick("startdate")  # still none

    # Final fallback: try any column that normalizes to "startdate"
    if date_col is None and "startdate" in cols:
        date_col = cols["startdate"]

    ret_col = pick("retailer", "store", "channel")

    if not sku_col or not price_col or not date_col:
        raise ValueError("Price history upload must include columns for SKU, Price, and StartDate (e.g., 'Start Date').")

    norm = pd.DataFrame({
        "SKU": n[sku_col].map(_normalize_sku),
        "Price": pd.to_numeric(n[price_col], errors="coerce"),
        "StartDate": pd.to_datetime(n[date_col], errors="coerce"),
        "Retailer": n[ret_col].map(_normalize_price_retailer) if ret_col else "*",
    })

    ignored = norm.copy()
    ignored["IgnoreReason"] = ""

    mask = ignored["SKU"].isna() | (ignored["SKU"].astype(str).str.strip() == "")
    ignored.loc[mask, "IgnoreReason"] = "Missing SKU"

    mask = ignored["StartDate"].isna()
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Missing StartDate",
        ignored.loc[mask, "IgnoreReason"]
    )

    mask = ignored["Price"].isna()
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Blank Price",
        ignored.loc[mask, "IgnoreReason"]
    )

    mask = (ignored["Price"].notna()) & (ignored["Price"] <= 0)
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Price <= 0",
        ignored.loc[mask, "IgnoreReason"]
    )

    keep = norm.dropna(subset=["SKU","StartDate","Price"]).copy()
    keep = keep[keep["Price"] > 0].copy()

    ignored = ignored[ignored["IgnoreReason"] != ""].copy()
    keep = keep.reset_index(drop=True)
    ignored = ignored.reset_index(drop=True)
    return keep, ignored


def _price_history_diff(cur: pd.DataFrame, incoming: pd.DataFrame) -> pd.DataFrame:
    """
    Build a diff table for (Retailer, SKU, StartDate).
    Actions: insert/update/noop
    """
    if cur is None or cur.empty:
        base = incoming.copy()
        base["OldPrice"] = np.nan
        base["Action"] = "insert"
        base["PriceDiff"] = np.nan
        return base[["Retailer","SKU","StartDate","OldPrice","Price","PriceDiff","Action"]].sort_values(["Retailer","SKU","StartDate"])

    cur2 = cur.copy()
    cur2["StartDate"] = pd.to_datetime(cur2["StartDate"], errors="coerce")
    inc = incoming.copy()
    inc["StartDate"] = pd.to_datetime(inc["StartDate"], errors="coerce")

    key = ["Retailer","SKU","StartDate"]
    merged = inc.merge(cur2[key + ["Price"]].rename(columns={"Price":"OldPrice"}), on=key, how="left")
    merged["PriceDiff"] = merged["Price"] - merged["OldPrice"]
    merged["Action"] = np.where(merged["OldPrice"].isna(), "insert",
                        np.where(np.isclose(merged["Price"], merged["OldPrice"], equal_nan=True), "noop", "update"))
    return merged[key + ["OldPrice","Price","PriceDiff","Action"]].sort_values(key)

def upsert_price_history(new_rows: pd.DataFrame) -> tuple[int, int, int]:
    """
    Upsert price history with effective dates.
    Returns (inserted, updated, ignored_noop) counts for reporting.
    """
    cur = load_price_history()
    incoming, _ignored = _prepare_price_history_upload(new_rows)

    if incoming.empty:
        return (0, 0, 0)

    diff = _price_history_diff(cur, incoming)
    to_apply = diff[diff["Action"].isin(["insert","update"])].copy()
    noop = int((diff["Action"] == "noop").sum())

    if to_apply.empty:
        return (0, 0, noop)

    apply_rows = to_apply[["Retailer","SKU","StartDate","Price"]].copy()

    merged = pd.concat([cur, apply_rows], ignore_index=True) if (cur is not None and not cur.empty) else apply_rows.copy()
    merged["StartDate"] = pd.to_datetime(merged["StartDate"], errors="coerce")
    merged = merged.dropna(subset=["SKU","Price","StartDate"])
    merged = merged.drop_duplicates(subset=["Retailer","SKU","StartDate"], keep="last")
    merged = merged.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
    save_price_history(merged)

    inserted = int((diff["Action"] == "insert").sum())
    updated = int((diff["Action"] == "update").sum())
    return (inserted, updated, noop)



def apply_effective_prices(base: pd.DataFrame, vmap: pd.DataFrame, ph: pd.DataFrame) -> pd.DataFrame:
    """
    Hybrid pricing:
      1) If UnitPrice is provided on the weekly sheet, ALWAYS use it (locks history).
      2) Else, use effective-date price history (retailer-specific first, then wildcard '*' retailer for all).
      3) Else, fall back to vendor map Price.

    Notes:
      - merge_asof requires non-null, sorted datetime keys.
    """
    base = base.copy()

    # Ensure expected columns exist
    if "Price" not in base.columns:
        base["Price"] = np.nan
    if "UnitPrice" not in base.columns:
        base["UnitPrice"] = np.nan

    base["StartDate"] = pd.to_datetime(base["StartDate"], errors="coerce")

    # Start with vendor-map price, then let weekly UnitPrice override
    base["PriceEffective"] = base["Price"]
    base["PriceEffective"] = base["UnitPrice"].combine_first(base["PriceEffective"])

    # If no price history, finish
    if ph is None or ph.empty:
        return base

    ph = ph.copy()
    ph["StartDate"] = pd.to_datetime(ph["StartDate"], errors="coerce")
    ph = ph.dropna(subset=["SKU", "StartDate", "Price"]).copy()
    if ph.empty:
        return base

    # Normalize keys
    if "Retailer" not in ph.columns:
        ph["Retailer"] = "*"
    ph["Retailer"] = ph["Retailer"].fillna("*").astype(str).str.strip()
    ph["SKU"] = ph["SKU"].map(_normalize_sku)
    base["SKU"] = base["SKU"].map(_normalize_sku)

    # merge_asof cannot handle NaT in the 'on' key
    base_valid = base[base["StartDate"].notna()].copy()
    base_invalid = base[base["StartDate"].isna()].copy()

    # Retailer-specific history (not '*')
    ph_exact = ph[ph["Retailer"] != "*"].copy()
    ph_star = ph[ph["Retailer"] == "*"].copy()

    # Apply retailer-specific prices
    if not ph_exact.empty and not base_valid.empty:
        b1 = base_valid.sort_values(["StartDate","Retailer","SKU"], kind="mergesort").reset_index(drop=True)
        p1 = ph_exact.sort_values(["StartDate","Retailer","SKU"], kind="mergesort").reset_index(drop=True)

        exact = pd.merge_asof(
            b1,
            p1[["Retailer", "SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_Price"}),
            by=["Retailer", "SKU"],
            on="StartDate",
            direction="backward",
            allow_exact_matches=True,
        )

        # Only use PH_Price when UnitPrice is missing
        exact["PriceEffective"] = exact["UnitPrice"].combine_first(exact["PH_Price"]).combine_first(exact["PriceEffective"])
        exact = exact.drop(columns=["PH_Price"], errors="ignore")
        base_valid = exact

    # Apply wildcard prices to rows still missing PriceEffective (and no UnitPrice)
    if not ph_star.empty and not base_valid.empty:
        missing = base_valid["UnitPrice"].isna() & base_valid["PriceEffective"].isna()
        if missing.any():
            b2 = base_valid.loc[missing].copy()
            b2 = b2.sort_values(["StartDate","SKU"], kind="mergesort").reset_index(drop=True)
            p2 = ph_star.sort_values(["StartDate","SKU"], kind="mergesort").reset_index(drop=True)

            star = pd.merge_asof(
                b2,
                p2[["SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_PriceStar"}),
                by=["SKU"],
                on="StartDate",
                direction="backward",
                allow_exact_matches=True,
            )
            base_valid.loc[missing, "PriceEffective"] = star["PH_PriceStar"].values

    # Final: ensure UnitPrice still wins
    if not base_valid.empty:
        base_valid["PriceEffective"] = base_valid["UnitPrice"].combine_first(base_valid["PriceEffective"])

    # Recombine
    base_out = pd.concat([base_valid, base_invalid], ignore_index=True)
    return base_out

    ph = ph.copy()
    ph["StartDate"] = pd.to_datetime(ph["StartDate"], errors="coerce")
    ph = ph.dropna(subset=["SKU", "StartDate", "Price"]).copy()

    if ph.empty:
        return base

    # Normalize retailer field
    if "Retailer" not in ph.columns:
        ph["Retailer"] = "*"
    ph["Retailer"] = ph["Retailer"].fillna("*").astype(str).str.strip()
    ph["SKU"] = ph["SKU"].map(_normalize_sku)
    base["SKU"] = base["SKU"].map(_normalize_sku)

    # Retailer-specific history (not '*')
    ph_exact = ph[ph["Retailer"] != "*"].copy()
    # Wildcard history applies to all retailers
    ph_star = ph[ph["Retailer"] == "*"].copy()

    # Apply retailer-specific prices using merge_asof
    if not ph_exact.empty:
        b1 = base.sort_values(["Retailer", "SKU", "StartDate"]).reset_index(drop=True)
        p1 = ph_exact.sort_values(["Retailer", "SKU", "StartDate"]).reset_index(drop=True)

        # merge_asof requires both sides sorted by by-keys then on-key
        exact = pd.merge_asof(
            b1,
            p1[["Retailer", "SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_Price"}),
            by=["Retailer", "SKU"],
            on="StartDate",
            direction="backward",
            allow_exact_matches=True,
        )
        base = exact

        # Only use PH_Price when UnitPrice is missing
        base["PriceEffective"] = base["UnitPrice"].combine_first(base["PH_Price"]).combine_first(base["PriceEffective"])
        base = base.drop(columns=["PH_Price"], errors="ignore")

    # Apply wildcard prices for any rows still missing an effective price (and no UnitPrice)
    if not ph_star.empty:
        missing = base["UnitPrice"].isna() & base["PriceEffective"].isna()
        if missing.any():
            b2 = base.loc[missing].copy()
            b2 = b2.sort_values(["SKU", "StartDate"]).reset_index(drop=True)
            p2 = ph_star.sort_values(["SKU", "StartDate"]).reset_index(drop=True)

            star = pd.merge_asof(
                b2,
                p2[["SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_PriceStar"}),
                by=["SKU"],
                on="StartDate",
                direction="backward",
                allow_exact_matches=True,
            )
            base.loc[missing, "PriceEffective"] = star["PH_PriceStar"].values

    # Final: ensure UnitPrice still wins
    base["PriceEffective"] = base["UnitPrice"].combine_first(base["PriceEffective"])

    return base

def upsert_sales(existing: pd.DataFrame, new_rows: pd.DataFrame) -> pd.DataFrame:
    if existing is None or existing.empty:
        return new_rows.copy()
    if new_rows is None or new_rows.empty:
        return existing.copy()

    for c in ["StartDate","EndDate"]:
        if c in existing.columns:
            existing[c] = pd.to_datetime(existing[c], errors="coerce")
        if c in new_rows.columns:
            new_rows[c] = pd.to_datetime(new_rows[c], errors="coerce")

    key_cols = ["Retailer","SKU","StartDate","EndDate","SourceFile"]
    combined = pd.concat([existing, new_rows], ignore_index=True)
    combined = combined.drop_duplicates(subset=key_cols, keep="last")
    return combined

def append_sales_to_store(new_rows: pd.DataFrame) -> None:
    if new_rows is None or new_rows.empty:
        return
    existing = load_sales_store()
    combined = upsert_sales(existing, new_rows)
    combined.to_csv(DEFAULT_SALES_STORE, index=False)

# -------------------------
# Weekly workbook ingestion
# -------------------------
def parse_date_range_from_filename(name: str, year_hint: int):
    n = name.lower()

    m = re.search(r"(\d{4})[-_/](\d{1,2})[-_/](\d{1,2}).*?(?:thru|through|to|–|-).*?(\d{4})[-_/](\d{1,2})[-_/](\d{1,2})", n)
    if m:
        y1, mo1, d1, y2, mo2, d2 = map(int, m.groups())
        return pd.Timestamp(date(y1, mo1, d1)), pd.Timestamp(date(y2, mo2, d2))

    m = re.search(r"(\d{1,2})[-_/](\d{1,2}).*?(?:thru|through|to|–|-).*?(\d{1,2})[-_/](\d{1,2})", n)
    if m:
        mo1, d1, mo2, d2 = map(int, m.groups())
        y = int(year_hint)
        return pd.Timestamp(date(y, mo1, d1)), pd.Timestamp(date(y, mo2, d2))

    return None, None

def read_weekly_workbook(uploaded_file, year: int) -> pd.DataFrame:
    xls = pd.ExcelFile(uploaded_file)
    fname = getattr(uploaded_file, "name", "upload.xlsx")
    sdt, edt = parse_date_range_from_filename(fname, year_hint=year)
    if sdt is None:
        sdt = pd.Timestamp(date.today() - timedelta(days=7))
        edt = pd.Timestamp(date.today())

    rows = []
    for sh in xls.sheet_names:
        retailer = _normalize_retailer(sh)
        raw = pd.read_excel(xls, sheet_name=sh, header=None)
        if raw.shape[1] < 2:
            continue
        raw = raw.iloc[:, :3].copy() if raw.shape[1] >= 3 else raw.iloc[:, :2].copy()
        raw.columns = ["SKU","Units","UnitPrice"] if raw.shape[1] == 3 else ["SKU","Units"]
        raw["SKU"] = raw["SKU"].map(_normalize_sku)
        raw["Units"] = pd.to_numeric(raw["Units"], errors="coerce").fillna(0.0)
        if "UnitPrice" in raw.columns:
            raw["UnitPrice"] = pd.to_numeric(raw["UnitPrice"], errors="coerce")
        else:
            raw["UnitPrice"] = np.nan
        raw = raw[raw["SKU"].astype(str).str.strip().ne("")]

        for _, r in raw.iterrows():
            rows.append({
                "Retailer": retailer,
                "SKU": r["SKU"],
                "Units": float(r["Units"]),
                "UnitPrice": float(r["UnitPrice"]) if pd.notna(r.get("UnitPrice", np.nan)) else np.nan,
                "StartDate": pd.to_datetime(sdt),
                "EndDate": pd.to_datetime(edt),
                "SourceFile": fname,
            })

    out = pd.DataFrame(rows)
    if not out.empty:
        out["Retailer"] = out["Retailer"].map(_normalize_retailer)
        out["SKU"] = out["SKU"].map(_normalize_sku)
        out["StartDate"] = pd.to_datetime(out["StartDate"], errors="coerce")
        out["EndDate"] = pd.to_datetime(out["EndDate"], errors="coerce")
    return out


# -------------------------
# Year-Overview (YOW) workbook ingestion
# -------------------------
def parse_week_range_header(val, year: int):
    """Parse headers like '1-1 / 1-3' into (StartDate, EndDate) timestamps.
    Accepts a few common variants and handles year-crossing weeks (e.g. '12-29 / 1-2').
    """
    if val is None:
        return (None, None)
    s = str(val).strip()
    if s == "":
        return (None, None)

    # Common: 'M-D / M-D' or 'M/D - M/D'
    m = re.search(r"(\d{1,2})\s*[-/]\s*(\d{1,2})\s*(?:/|to|–|-)+\s*(\d{1,2})\s*[-/]\s*(\d{1,2})", s)
    if not m:
        # Variant with explicit months: '1-1 / 1-3' (same as above but stricter)
        m = re.search(r"(\d{1,2})-(\d{1,2})\s*/\s*(\d{1,2})-(\d{1,2})", s)
    if not m:
        return (None, None)

    mo1, d1, mo2, d2 = map(int, m.groups())
    y1 = int(year)
    y2 = int(year)
    # If the end month is earlier than start month, assume it crosses into next year
    if mo2 < mo1:
        y2 = y1 + 1

    try:
        sdt = pd.Timestamp(y1, mo1, d1)
        edt = pd.Timestamp(y2, mo2, d2)
        return (sdt, edt)
    except Exception:
        return (None, None)

def read_yow_workbook(uploaded_file, year: int) -> pd.DataFrame:
    """Read a Year Overview workbook:
    - One sheet per retailer OR a single sheet where A1 is the retailer name.
    - Row 1 contains week ranges across the top (starting in column B).
    - Column A contains SKUs (starting row 2).
    - Cells contain Units.
    """
    import openpyxl

    fname = getattr(uploaded_file, "name", "yow.xlsx")

    # openpyxl is fastest/most tolerant for wide sheets
    wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True, keep_links=False)

    rows_out = []

    for sh in wb.sheetnames:
        ws = wb[sh]

        # Retailer name: A1 (preferred). If blank, fall back to sheet name.
        retailer_name = ws["A1"].value
        retailer = _normalize_retailer(retailer_name if retailer_name not in [None, ""] else sh)

        # Header row: week ranges from B1 onward until blank
        week_cols = []
        col = 2  # B
        while True:
            v = ws.cell(row=1, column=col).value
            if v is None or str(v).strip() == "":
                break
            sdt, edt = parse_week_range_header(v, year=year)
            if sdt is None:
                # Try interpreting as a date (week start) if someone uses real date headers
                dt = pd.to_datetime(v, errors="coerce")
                if pd.notna(dt):
                    sdt = pd.Timestamp(dt).normalize()
                    edt = sdt + pd.Timedelta(days=6)
                else:
                    # stop if header isn't parseable
                    break
            if edt is None:
                edt = sdt + pd.Timedelta(days=6)
            week_cols.append((col, pd.Timestamp(sdt), pd.Timestamp(edt), str(v).strip()))
            col += 1

        if not week_cols:
            continue

        # Data rows: SKUs down column A from row 2 until blank
        row = 2
        while True:
            sku = ws.cell(row=row, column=1).value
            if sku is None or str(sku).strip() == "":
                break
            sku = _normalize_sku(sku)

            for (cidx, sdt, edt, hdr) in week_cols:
                units = ws.cell(row=row, column=cidx).value
                if units is None or (isinstance(units, str) and units.strip() == ""):
                    continue
                try:
                    u = float(units)
                except Exception:
                    continue
                if np.isnan(u) or u == 0:
                    continue

                rows_out.append({
                    "Retailer": retailer,
                    "SKU": sku,
                    "Units": float(u),
                    "UnitPrice": np.nan,          # use current pricing (vendor map / price history)
                    "StartDate": pd.to_datetime(sdt),
                    "EndDate": pd.to_datetime(edt),
                    "SourceFile": f"{fname}::{sh}",
                })

            row += 1

    out = pd.DataFrame(rows_out)
    if not out.empty:
        out["Retailer"] = out["Retailer"].map(_normalize_retailer)
        out["SKU"] = out["SKU"].map(_normalize_sku)
        out["StartDate"] = pd.to_datetime(out["StartDate"], errors="coerce")
        out["EndDate"] = pd.to_datetime(out["EndDate"], errors="coerce")
        out["Units"] = pd.to_numeric(out["Units"], errors="coerce").fillna(0.0)
        out["UnitPrice"] = pd.to_numeric(out["UnitPrice"], errors="coerce")
    return out

# -------------------------
# Enrichment / metrics
# -------------------------
def enrich_sales(sales: pd.DataFrame, vmap: pd.DataFrame, price_hist: pd.DataFrame | None = None) -> pd.DataFrame:
    s = sales.copy()
    s["Retailer"] = s["Retailer"].map(_normalize_retailer)
    s["SKU"] = s["SKU"].map(_normalize_sku)
    s["Units"] = pd.to_numeric(s["Units"], errors="coerce").fillna(0.0).astype(float)
    s["StartDate"] = pd.to_datetime(s["StartDate"], errors="coerce")
    s["EndDate"] = pd.to_datetime(s["EndDate"], errors="coerce")

    m = vmap[["Retailer","SKU","Vendor","Price","MapOrder"]].copy()
    m["Retailer"] = m["Retailer"].map(_normalize_retailer)
    m["SKU"] = m["SKU"].map(_normalize_sku)
    m["Price"] = pd.to_numeric(m["Price"], errors="coerce")

    out = s.merge(m, on=["Retailer","SKU"], how="left")

    # Apply effective-dated pricing (if provided), otherwise fallback to vendor map price
    ph = price_hist if price_hist is not None else load_price_history()
    out = apply_effective_prices(out, vmap, ph)


    # Compute Sales from Units and effective price (Units-only weekly uploads)
    out["Units"] = pd.to_numeric(out.get("Units", 0), errors="coerce").fillna(0.0)
    out["PriceEffective"] = pd.to_numeric(out.get("PriceEffective", np.nan), errors="coerce")
    out["Sales"] = (out["Units"] * out["PriceEffective"]).fillna(0.0)
    return out

def wow_mom_metrics(df: pd.DataFrame) -> dict:
    out = {"total_units":0.0,"total_sales":0.0,"wow_units":None,"wow_sales":None,"mom_units":None,"mom_sales":None}
    if df is None or df.empty:
        return out
    d = df.copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    out["total_units"] = float(d["Units"].sum())
    out["total_sales"] = float(d["Sales"].fillna(0).sum())

    periods = sorted(d["StartDate"].dropna().dt.date.unique().tolist())
    if len(periods) >= 1:
        cur_p = periods[-1]
        cur = d[d["StartDate"].dt.date == cur_p]
        cur_u = cur["Units"].sum()
        cur_s = cur["Sales"].fillna(0).sum()
        if len(periods) >= 2:
            prev_p = periods[-2]
            prev = d[d["StartDate"].dt.date == prev_p]
            prev_u = prev["Units"].sum()
            prev_s = prev["Sales"].fillna(0).sum()
        else:
            prev_u = 0.0
            prev_s = 0.0
        out["wow_units"] = float(cur_u - prev_u)
        out["wow_sales"] = float(cur_s - prev_s)

    d["MonthP"] = d["StartDate"].dt.to_period("M")
    months = sorted(d["MonthP"].dropna().unique().tolist())
    if len(months) >= 1:
        cur_m = months[-1]
        cur = d[d["MonthP"] == cur_m]
        cur_u = cur["Units"].sum()
        cur_s = cur["Sales"].fillna(0).sum()
        if len(months) >= 2:
            prev_m = months[-2]
            prev = d[d["MonthP"] == prev_m]
            prev_u = prev["Units"].sum()
            prev_s = prev["Sales"].fillna(0).sum()
        else:
            prev_u = 0.0
            prev_s = 0.0
        out["mom_units"] = float(cur_u - prev_u)
        out["mom_sales"] = float(cur_s - prev_s)

    return out

def month_label(p: pd.Period) -> str:
    return p.to_timestamp().strftime("%B %Y")

# -------------------------
# App UI
# -------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

with st.sidebar:
    st.header("Data Inputs")
    edit_mode = st.checkbox("Enable Edit Mode (edit Vendor/Price)", value=False)

    this_year = date.today().year
    year = st.selectbox("Year (for filename date parsing)", options=list(range(this_year-3, this_year+2)), index=3)
    view_year = st.selectbox("View Year (dashboard)", options=list(range(this_year-3, this_year+2)), index=3, key="view_year")

    st.subheader("Vendor Map")
    vm_upload = st.file_uploader("Upload Vendor Map (.xlsx)", type=["xlsx"], key="vm_up")
    a, b = st.columns(2)
    with a:
        if st.button("Use uploaded as default", disabled=vm_upload is None):
            DEFAULT_VENDOR_MAP.write_bytes(vm_upload.getbuffer())
            st.success("Saved as default vendor map.")
            st.rerun()
    with b:
        if st.button("Reload"):
            st.rerun()

    st.subheader("Weekly Sales Workbooks")
    wk_uploads = st.file_uploader("Upload weekly sales workbook(s) (.xlsx)", type=["xlsx"], accept_multiple_files=True, key="wk_up")
    if st.button("Ingest uploads", disabled=not wk_uploads):
        for f in wk_uploads:
            new_rows = read_weekly_workbook(f, year=year)
            append_sales_to_store(new_rows)
        st.success("Ingested uploads into the sales store.")
        st.rerun()

    st.divider()
    if st.button("Clear ALL stored sales data"):
        if DEFAULT_SALES_STORE.exists():
            DEFAULT_SALES_STORE.unlink()
        st.warning("Sales store cleared.")
        st.rerun()


# Ensure view_year exists for downstream tabs
view_year = st.session_state.get('view_year', year)

# Load vendor map
if vm_upload is not None:
    tmp = DATA_DIR / "_session_vendor_map.xlsx"
    tmp.write_bytes(vm_upload.getbuffer())
    vmap = load_vendor_map(tmp)
elif DEFAULT_VENDOR_MAP.exists():
    vmap = load_vendor_map(DEFAULT_VENDOR_MAP)
else:
    st.info("Upload a vendor map to begin.")
    st.stop()

sales_store = load_sales_store()
price_hist = load_price_history()
df_all = enrich_sales(sales_store, vmap, price_hist)

# KPIs across top (always current calendar year)
df_kpi = df_all.copy()
df_kpi["StartDate"] = pd.to_datetime(df_kpi["StartDate"], errors="coerce")
df_kpi = df_kpi[df_kpi["StartDate"].dt.year == int(this_year)].copy()

# Apply view-year filter for all reporting tabs
df = df_all.copy()
df["StartDate"] = pd.to_datetime(df["StartDate"], errors="coerce")
df = df[df["StartDate"].dt.year == int(view_year)].copy()

# KPIs across top
m_all = wow_mom_metrics(df_kpi)

st.markdown("## 📊 Overview (All Retailers)")
r1 = st.columns(3)
r2 = st.columns(3)
with r1[0]:
    st.metric("Total Units (YTD)", fmt_int(m_all["total_units"]))
with r1[1]:
    st.metric("Total Sales (YTD)", fmt_currency(m_all["total_sales"]))
with r1[2]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>MoM Units</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['mom_units'])};'>{fmt_int(m_all['mom_units']) if m_all['mom_units'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )
with r2[0]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>MoM Sales</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['mom_sales'])};'>{fmt_currency(m_all['mom_sales']) if m_all['mom_sales'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )
with r2[1]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>WoW Units</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['wow_units'])};'>{fmt_int(m_all['wow_units']) if m_all['wow_units'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )
with r2[2]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>WoW Sales</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['wow_sales'])};'>{fmt_currency(m_all['wow_sales']) if m_all['wow_sales'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )

st.divider()


# -------------------------
# Reporting helpers
# -------------------------
def week_labels(df_in: pd.DataFrame) -> list[str]:
    if df_in is None or df_in.empty:
        return []
    w = sorted(pd.to_datetime(df_in["StartDate"], errors="coerce").dropna().dt.date.unique().tolist())
    return [pd.Timestamp(x).strftime("%m-%d") for x in w]

def add_week_col(d: pd.DataFrame) -> pd.DataFrame:
    d = d.copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    d["Week"] = d["StartDate"].dt.date
    return d

def nonzero_mean_rowwise(frame: pd.DataFrame) -> pd.Series:
    """Mean across columns, ignoring zeros (treat zeros as missing)."""
    return frame.replace(0, np.nan).mean(axis=1)

def last_n_weeks(df_in: pd.DataFrame, n: int):
    if df_in is None or df_in.empty:
        return []
    w = sorted(pd.to_datetime(df_in["StartDate"], errors="coerce").dropna().dt.date.unique().tolist())
    return w[-n:] if len(w) >= n else w

def safe_div(a, b):
    try:
        if b == 0 or pd.isna(b):
            return np.nan
        return a / b
    except Exception:
        return np.nan

def to_pdf_bytes(title: str, sections: list[tuple[str, list[str]]]) -> bytes:
    """
    Build a simple PDF summary.
    sections: list of (heading, lines[])
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
    except Exception:
        return b""

    import io
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter
    x = 0.75 * inch
    y = height - 0.75 * inch

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x, y, title)
    y -= 0.35 * inch

    for heading, lines in sections:
        if y < 1.0 * inch:
            c.showPage()
            y = height - 0.75 * inch
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x, y, heading)
        y -= 0.22 * inch
        c.setFont("Helvetica", 10)
        for ln in lines:
            if y < 1.0 * inch:
                c.showPage()
                y = height - 0.75 * inch
                c.setFont("Helvetica", 10)
            c.drawString(x, y, str(ln)[:120])
            y -= 0.18 * inch
        y -= 0.10 * inch

    c.save()
    return buf.getvalue()

tabs = st.tabs([
    "Retailer Totals",
    "Vendor Totals",
    "Unit Summary",
    "Executive Summary",
    "WoW Exceptions",
    "Comparison",
    "Year Summary",
    "Data Inventory",
    "Insights & Alerts",
    "Run-Rate Forecast",
    "No Sales SKUs",
    "Edit Vendor Map",
    "Backup / Restore",
    "Bulk Data Upload",
])
(tab_retail_totals, tab_vendor_totals, tab_unit_summary, tab_exec, tab_wow_exc,
 tab_compare, tab_year_summary, tab_inventory, tab_alerts, tab_runrate, tab_no_sales,
 tab_edit_map, tab_backup, tab_bulk_upload) = tabs








def resolve_week_dates(periods: list, window):
    """
    periods: sorted list of datetime.date representing week start dates.
    window: int weeks or string like "6 months".
    Returns list of week dates to include, ordered ascending.
    """
    if not periods:
        return []
    if isinstance(window, int):
        return periods[-window:] if len(periods) >= window else periods
    if isinstance(window, str) and "month" in window:
        try:
            n = int(window.split()[0])
        except Exception:
            n = 6
        # get last n unique months present in periods
        months = [pd.Timestamp(d).to_period("M") for d in periods]
        uniq = []
        for p in months:
            if p not in uniq:
                uniq.append(p)
        usem = uniq[-n:] if len(uniq) >= n else uniq
        use = [d for d in periods if pd.Timestamp(d).to_period("M") in usem]
        return use
    return periods


def make_totals_tables(base: pd.DataFrame, group_col: str, tf_weeks, avg_weeks):
    if base.empty:
        return pd.DataFrame(), pd.DataFrame()
    base = base.copy()
    base["StartDate"] = pd.to_datetime(base["StartDate"], errors="coerce")
    periods = sorted(base["StartDate"].dropna().dt.date.unique().tolist())
    first_week = periods[0] if periods else None
    if not periods:
        return pd.DataFrame(), pd.DataFrame()

    use = resolve_week_dates(periods, tf_weeks)
    d = base[base["StartDate"].dt.date.isin(use)].copy()
    d["Week"] = d["StartDate"].dt.date

    sales_p = d.pivot_table(index=group_col, columns="Week", values="Sales", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)
    units_p = d.pivot_table(index=group_col, columns="Week", values="Units", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)

    if len(use) >= 2:
        sales_p["Diff"] = sales_p[use[-1]] - sales_p[use[-2]]
        units_p["Diff"] = units_p[use[-1]] - units_p[use[-2]]
    else:
        sales_p["Diff"] = 0.0
        units_p["Diff"] = 0.0

    # Determine which weeks to average based on selected average window
    current_year = int(pd.to_datetime(base["StartDate"], errors="coerce").dt.year.max() or date.today().year)
    avg_use = resolve_avg_use(avg_weeks, use, current_year)

    # Ignore the very first week of the year (partial week)
    if first_week is not None and avg_use:
        avg_use = [w for w in avg_use if pd.to_datetime(w, errors="coerce").date() != first_week]

    sales_p["Avg"] = sales_p[avg_use].replace(0, np.nan).mean(axis=1) if avg_use else 0.0
    units_p["Avg"] = units_p[avg_use].replace(0, np.nan).mean(axis=1) if avg_use else 0.0

    # Diff vs Avg uses the last week displayed minus Avg
    if use:
        sales_p["Diff vs Avg"] = sales_p[use[-1]] - sales_p["Avg"]
        units_p["Diff vs Avg"] = units_p[use[-1]] - units_p["Avg"]
    else:
        sales_p["Diff vs Avg"] = 0.0
        units_p["Diff vs Avg"] = 0.0

    sales_p = sales_p.sort_index()
    units_p = units_p.sort_index()

    sales_p.loc["TOTAL"] = sales_p.sum(axis=0)
    units_p.loc["TOTAL"] = units_p.sum(axis=0)

    # Recompute TOTAL Avg and Diff vs Avg from totals row values
    if "Avg" in sales_p.columns and use:
        sales_p.loc["TOTAL","Avg"] = sales_p.loc["TOTAL", [c for c in avg_use]].replace(0, np.nan).mean() if avg_use else 0.0
        units_p.loc["TOTAL","Avg"] = units_p.loc["TOTAL", [c for c in avg_use]].replace(0, np.nan).mean() if avg_use else 0.0
        sales_p.loc["TOTAL","Diff vs Avg"] = sales_p.loc["TOTAL", use[-1]] - sales_p.loc["TOTAL","Avg"]
        units_p.loc["TOTAL","Diff vs Avg"] = units_p.loc["TOTAL", use[-1]] - units_p.loc["TOTAL","Avg"]

    def wlab(c):
        try:
            return pd.Timestamp(c).strftime("%m-%d")
        except Exception:
            return c

    sales_p = sales_p.rename(columns={c: wlab(c) for c in sales_p.columns})
    units_p = units_p.rename(columns={c: wlab(c) for c in units_p.columns})

    return sales_p.reset_index(), units_p.reset_index()

# Retailer Totals
with tab_retail_totals:
    st.subheader("Retailer Totals")
    tf = st.selectbox("Timeframe", options=[2,4,8,12,"4 months", "5 months", "6 months", "7 months", "8 months", "9 months", "10 months", "11 months", "12 months"], index=1, key="rt_tf")
    avgw = st.selectbox("Average window", options=AVG_WINDOW_OPTIONS, index=0, key="rt_avg")

    sales_t, units_t = make_totals_tables(df, "Retailer", tf, avgw)
    if sales_t.empty:
        st.info("No data yet.")
    else:
        st.markdown("### Sales ($) by Week")
        st.dataframe(style_currency_cols(sales_t, diff_cols=["Diff","Diff vs Avg"]), use_container_width=True, height=_table_height(sales_t), hide_index=True)

        st.markdown("### Units by Week")
        ud = units_t.copy()
        first = "Retailer"
        for c in ud.columns:
            if c == first:
                continue
            if c == "Avg":
                ud[c] = ud[c].astype(float)
            else:
                ud[c] = ud[c].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
        sty = ud.style
        if "Diff" in ud.columns:
            sty = sty.applymap(lambda v: f"color: {_color(v)};", subset=["Diff","Diff vs Avg"])
        fmt = {}
        for c in ud.columns:
            if c == first:
                continue
            fmt[c] = (lambda v: fmt_2(v)) if c == "Avg" else (lambda v: fmt_int(v))
        sty = sty.format(fmt)
        st.dataframe(sty, use_container_width=True, height=_table_height(ud), hide_index=True)

# Vendor Totals
with tab_vendor_totals:
    st.subheader("Vendor Totals")
    tf = st.selectbox("Timeframe", options=[2,4,8,12,"4 months", "5 months", "6 months", "7 months", "8 months", "9 months", "10 months", "11 months", "12 months"], index=1, key="vt_tf")
    avgw = st.selectbox("Average window", options=AVG_WINDOW_OPTIONS, index=0, key="vt_avg")

    base = df.copy()
    base["Vendor"] = base["Vendor"].fillna("Unmapped")
    sales_t, units_t = make_totals_tables(base, "Vendor", tf, avgw)
    if sales_t.empty:
        st.info("No data yet.")
    else:
        st.markdown("### Sales ($) by Week")
        st.dataframe(style_currency_cols(sales_t, diff_cols=["Diff","Diff vs Avg"]), use_container_width=True, height=_table_height(sales_t, max_px=1400), hide_index=True)

        st.markdown("### Units by Week")
        ud = units_t.copy()
        first = "Vendor"
        for c in ud.columns:
            if c == first:
                continue
            if c == "Avg":
                ud[c] = ud[c].astype(float)
            else:
                ud[c] = ud[c].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
        sty = ud.style
        if "Diff" in ud.columns:
            sty = sty.applymap(lambda v: f"color: {_color(v)};", subset=["Diff","Diff vs Avg"])
        fmt = {}
        for c in ud.columns:
            if c == first:
                continue
            fmt[c] = (lambda v: fmt_2(v)) if c == "Avg" else (lambda v: fmt_int(v))
        sty = sty.format(fmt)
        st.dataframe(sty, use_container_width=True, height=_table_height(ud, max_px=1400), hide_index=True)

# Unit Summary
with tab_unit_summary:
    st.subheader("Unit Summary")
    retailers = sorted(vmap["Retailer"].dropna().unique().tolist())
    sel_r = st.selectbox("Retailer", options=retailers, index=0, key="us_retailer")
    tf = st.selectbox("Timeframe", options=[2,4,8,12,"4 months", "5 months", "6 months", "7 months", "8 months", "9 months", "10 months", "11 months", "12 months"], index=1, key="us_tf")
    avgw = st.selectbox("Average window", options=AVG_WINDOW_OPTIONS, index=0, key="us_avg")

    d = df[df["Retailer"] == sel_r].copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    periods = sorted(d["StartDate"].dropna().dt.date.unique().tolist())
    first_week = periods[0] if periods else None
    use = resolve_week_dates(periods, tf)
    if not use:
        st.info("No data for this retailer yet.")
    else:
        d = d[d["StartDate"].dt.date.isin(use)].copy()
        d["Week"] = d["StartDate"].dt.date

        sku_order = vmap[vmap["Retailer"] == sel_r].sort_values("MapOrder")["SKU"].tolist()

        units_p = d.pivot_table(index="SKU", columns="Week", values="Units", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)
        units_p = units_p.loc[units_p.sum(axis=1) > 0]
        units_p = units_p.reindex([s for s in sku_order if s in units_p.index])

        if units_p.empty:
            st.info("No SKUs sold in this timeframe.")
        else:
            if len(use) >= 2:
                units_p["Diff"] = units_p[use[-1]] - units_p[use[-2]]
            else:
                units_p["Diff"] = 0.0
            avg_use = resolve_week_dates(use, avgw)
            if first_week is not None and avg_use:
                avg_use = [w for w in avg_use if w != first_week]
            units_p["Avg"] = units_p[avg_use].replace(0, np.nan).mean(axis=1) if avg_use else 0.0
            units_p["Diff vs Avg"] = units_p[use[-1]] - units_p["Avg"] if use else 0.0

            units_p = units_p.rename(columns={c: pd.Timestamp(c).strftime("%m-%d") for c in use})
            units_out = units_p.reset_index()
            total = {"SKU":"TOTAL"}
            for c in units_out.columns:
                if c != "SKU":
                    total[c] = float(units_out[c].sum())
            units_out = pd.concat([units_out, pd.DataFrame([total])], ignore_index=True)

            ud = units_out.copy()
            for c in ud.columns:
                if c == "SKU":
                    continue
                if c == "Avg":
                    ud[c] = ud[c].astype(float)
                else:
                    ud[c] = ud[c].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
            sty = ud.style
            if "Diff" in ud.columns:
                sty = sty.applymap(lambda v: f"color: {_color(v)};", subset=["Diff","Diff vs Avg"])
            fmt = {}
            for c in ud.columns:
                if c == "SKU":
                    continue
                fmt[c] = (lambda v: fmt_2(v)) if c == "Avg" else (lambda v: fmt_int(v))
            sty = sty.format(fmt)

            st.markdown("### Units by Week (per SKU)")
            st.dataframe(sty, use_container_width=True, height=_table_height(ud, max_px=1400), hide_index=True)

            sales_p = d.pivot_table(index="SKU", columns="Week", values="Sales", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)
            sales_p = sales_p.loc[sales_p.sum(axis=1) > 0]
            sales_p = sales_p.reindex([s for s in sku_order if s in sales_p.index])

            if len(use) >= 2:
                sales_p["Diff"] = sales_p[use[-1]] - sales_p[use[-2]]
            else:
                sales_p["Diff"] = 0.0
            avg_use = resolve_week_dates(use, avgw)
            if first_week is not None and avg_use:
                avg_use = [w for w in avg_use if w != first_week]
            sales_p["Avg"] = sales_p[avg_use].replace(0, np.nan).mean(axis=1) if avg_use else 0.0
            sales_p["Diff vs Avg"] = sales_p[use[-1]] - sales_p["Avg"] if use else 0.0

            sales_p = sales_p.rename(columns={c: pd.Timestamp(c).strftime("%m-%d") for c in use})
            sales_out = sales_p.reset_index()
            total = {"SKU":"TOTAL"}
            for c in sales_out.columns:
                if c != "SKU":
                    total[c] = float(sales_out[c].sum())
            sales_out = pd.concat([sales_out, pd.DataFrame([total])], ignore_index=True)

            st.markdown("### Sales ($) by Week (per SKU)")
            st.dataframe(style_currency_cols(sales_out, diff_cols=["Diff","Diff vs Avg"]), use_container_width=True, height=_table_height(sales_out, max_px=1400), hide_index=True)

def monthly_totals(d: pd.DataFrame):
    if d.empty:
        return pd.DataFrame(columns=["Month","Units","Sales"])
    d = d.copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    d["MonthP"] = d["StartDate"].dt.to_period("M")
    agg = d.groupby("MonthP", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
    agg["Month"] = agg["MonthP"].map(month_label)
    agg = agg.sort_values("MonthP")
    return agg[["Month","Units","Sales"]]

def score_kpis(d: pd.DataFrame):
    m = wow_mom_metrics(d)
    left, right = st.columns(2)
    with left:
        st.metric("YTD Units", fmt_int(m["total_units"]))
        st.markdown(
            f"<div style='font-size:13px; color: gray;'>WoW Units</div>"
            f"<div style='font-size:22px; font-weight:600; color:{_color(m['wow_units'])};'>{fmt_int(m['wow_units']) if m['wow_units'] is not None else '—'}</div>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<div style='font-size:13px; color: gray;'>MoM Units</div>"
            f"<div style='font-size:22px; font-weight:600; color:{_color(m['mom_units'])};'>{fmt_int(m['mom_units']) if m['mom_units'] is not None else '—'}</div>",
            unsafe_allow_html=True
        )
    with right:
        st.metric("YTD Sales", fmt_currency(m["total_sales"]))
        st.markdown(
            f"<div style='font-size:13px; color: gray;'>WoW Sales</div>"
            f"<div style='font-size:22px; font-weight:600; color:{_color(m['wow_sales'])};'>{fmt_currency(m['wow_sales']) if m['wow_sales'] is not None else '—'}</div>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<div style='font-size:13px; color: gray;'>MoM Sales</div>"
            f"<div style='font-size:22px; font-weight:600; color:{_color(m['mom_sales'])};'>{fmt_currency(m['mom_sales']) if m['mom_sales'] is not None else '—'}</div>",
            unsafe_allow_html=True
        )
    st.divider()


# No Sales SKUs
with tab_no_sales:
    st.subheader("No Sales SKUs")
    weeks = st.selectbox("Timeframe (weeks)", options=[3,6,8,12], index=0, key="ns_weeks")
    retailers = sorted(vmap["Retailer"].dropna().unique().tolist())
    sel_r = st.selectbox("Retailer", options=["All"] + retailers, index=0, key="ns_retailer")

    if df.empty:
        st.info("No sales data yet.")
    else:
        d2 = df.copy()
        d2["StartDate"] = pd.to_datetime(d2["StartDate"], errors="coerce")
        periods = sorted(d2["StartDate"].dropna().dt.date.unique().tolist())
        use = periods[-weeks:] if len(periods) >= weeks else periods

        if not use:
            st.info("No periods found yet.")
        else:
            sold = d2[d2["StartDate"].dt.date.isin(use)].groupby(["Retailer","SKU"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
            ref = vmap[["Retailer","SKU","Vendor","MapOrder"]].copy()
            if sel_r != "All":
                ref = ref[ref["Retailer"] == sel_r].copy()

            merged = ref.merge(sold, on=["Retailer","SKU"], how="left")
            merged["Units"] = merged["Units"].fillna(0.0)
            merged["Sales"] = merged["Sales"].fillna(0.0)

            nos = merged[(merged["Units"] <= 0) & (merged["Sales"] <= 0)].copy()
            nos["Status"] = f"No sales in last {weeks} weeks"
            nos = nos.sort_values(["Retailer","MapOrder","SKU"], ascending=[True, True, True])

            out = nos[["Retailer","Vendor","SKU","Status"]].copy()
            st.dataframe(out, use_container_width=True, height=_table_height(out, max_px=1400), hide_index=True)


# -------------------------
# WoW Exceptions
# -------------------------


with tab_wow_exc:
    st.subheader("Exceptions (Recent Week vs Prior Average)")

    if df.empty:
        st.info("No sales data yet.")
    else:
        d = add_week_col(df)
        weeks_all = sorted(d["Week"].dropna().unique().tolist())
        if len(weeks_all) < 2:
            st.info("Need at least two weeks of data to compute changes.")
        else:
            # N = number of prior weeks used to compute the baseline average (excluding the most recent week)
            n_prior = st.selectbox("Baseline average over prior window", options=["1 week","2 weeks","3 weeks","4 weeks","5 weeks","6 weeks","7 weeks","8 weeks","3 months","6 months"], index=3, key="wow_nprior")
            direction = st.selectbox("Direction", options=["Increase", "Decrease"], index=0, key="wow_dir")
            thresh = st.selectbox(
                "Percent threshold",
                options=[0.05,0.10,0.15,0.20,0.25,0.30,0.40,0.50,0.60,0.70,0.80,0.90,1.00,1.50],
                index=3,
                format_func=lambda x: f"{int(x*100)}%",
                key="wow_thresh2"
            )

            scope = st.selectbox("Scope", options=["All", "Retailer", "Vendor"], index=0, key="wow_scope2")
            if scope == "Retailer":
                opts = sorted(vmap["Retailer"].dropna().unique().tolist())
                pick = st.selectbox("Retailer", options=opts, index=0, key="wow_pick_r2")
                d2 = d[d["Retailer"] == pick].copy()
            elif scope == "Vendor":
                opts = sorted([v for v in vmap["Vendor"].dropna().unique().tolist() if str(v).strip() != ""])
                pick = st.selectbox("Vendor", options=opts, index=0, key="wow_pick_v2")
                d2 = d[d["Vendor"] == pick].copy()
            else:
                d2 = d

            weeks_all2 = sorted(d2["Week"].dropna().unique().tolist())
            if len(weeks_all2) < 2:
                st.info("Not enough weeks for this selection.")
            else:
                end_week = weeks_all2[-1]
                prior_weeks = weeks_all2[:-1]

                # Determine prior window based on selector (excluding most recent week)
                sel = n_prior
                if isinstance(sel, str) and "month" in sel:
                    nmo = int(sel.split()[0])
                    tmp = d2[d2["Week"].isin(prior_weeks)].copy()
                    tmp["MonthP"] = pd.to_datetime(tmp["StartDate"], errors="coerce").dt.to_period("M")
                    months = sorted(tmp["MonthP"].dropna().unique().tolist())
                    usem = months[-nmo:] if len(months) >= nmo else months
                    tmp = tmp[tmp["MonthP"].isin(usem)]
                    prior_weeks = sorted(tmp["Week"].dropna().unique().tolist())
                else:
                    nw = int(str(sel).split()[0])
                    prior_weeks = prior_weeks[-nw:] if len(prior_weeks) >= nw else prior_weeks

                weekly = d2.groupby(["Retailer","Vendor","SKU","Week"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))

                cur = weekly[weekly["Week"] == end_week].groupby(["Retailer","Vendor","SKU"], as_index=False).agg(
                    Units_End=("Units","sum"),
                    Sales_End=("Sales","sum")
                )

                if len(prior_weeks) == 0:
                    st.info("Need at least one prior week to build the baseline average.")
                else:
                    base = weekly[weekly["Week"].isin(prior_weeks)].groupby(["Retailer","Vendor","SKU","Week"], as_index=False).agg(
                        Units=("Units","sum"),
                        Sales=("Sales","sum")
                    )
                    base_avg = base.groupby(["Retailer","Vendor","SKU"], as_index=False).agg(
                        Units_Base=("Units","mean"),
                        Sales_Base=("Sales","mean")
                    )

                    res = cur.merge(base_avg, on=["Retailer","Vendor","SKU"], how="outer").fillna(0.0)
                    res["Units_Diff"] = res["Units_End"] - res["Units_Base"]
                    res["Sales_Diff"] = res["Sales_End"] - res["Sales_Base"]
                    res["Units_Pct"] = res["Units_Diff"] / res["Units_Base"].replace(0, np.nan)


                    # Require at least one prior week with sales/units (exclude brand-new items)
                    res = res[(res["Units_Base"] > 0) | (res["Sales_Base"] > 0)]
                    if direction == "Increase":
                        res = res[(res["Units_Pct"] >= thresh) & res["Units_Pct"].notna()]
                        res = res.sort_values(["Units_Diff","Sales_Diff"], ascending=[False, False])
                    else:
                        res = res[(res["Units_Pct"] <= -thresh) & res["Units_Pct"].notna()]
                        res = res.sort_values(["Units_Diff","Sales_Diff"], ascending=[True, True])

                    res = res.head(100)

                    if res.empty:
                        st.info("No items met the threshold for this selection.")
                    else:
                        t = res[["Retailer","Vendor","SKU","Units_Base","Units_End","Units_Diff","Units_Pct","Sales_Base","Sales_End","Sales_Diff"]].copy()
                        t["Units_Base"] = t["Units_Base"].map(lambda v: float(v))
                        t["Units_End"] = t["Units_End"].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
                        t["Units_Diff"] = t["Units_Diff"].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
                        # Rename for display
                        t = t.rename(columns={"Units_Pct":"% Diff"})

                        sty = t.style.format({
                            "Units_Base": lambda v: fmt_2(v),
                            "Units_End": lambda v: fmt_int(v),
                            "Units_Diff": lambda v: fmt_int(v),
                            "% Diff": lambda v: f"{(v*100):.1f}%" if pd.notna(v) else "—",
                            "Sales_Base": lambda v: fmt_currency(v),
                            "Sales_End": lambda v: fmt_currency(v),
                            "Sales_Diff": lambda v: fmt_currency(v),
                        }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])

                        st.caption(
                            f"Comparing most recent week ({pd.Timestamp(end_week).strftime('%m-%d')}) "
                            f"to the average of prior {len(prior_weeks)} week(s): "
                            + ", ".join([pd.Timestamp(w).strftime('%m-%d') for w in prior_weeks])
                        )
                        st.dataframe(sty, use_container_width=True, height=_table_height(t, max_px=1200), hide_index=True)




# -------------------------
# Comparison
# -------------------------
with tab_compare:
    st.subheader("Comparison (Month vs Month / Multi-month)")

    if df.empty:
        st.info("No sales data yet.")
    else:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()

        # Month options across ALL years (so you can compare Jan 2025 vs Jan 2026)
        d["MonthP"] = d["StartDate"].dt.to_period("M")

        months = sorted(d["MonthP"].unique().tolist())
        month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
        label_to_period = dict(zip(month_labels, months))

        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            a_pick = st.multiselect(
                "Selection A (one or more months)",
                options=month_labels,
                default=month_labels[-1:] if month_labels else [],
                key="cmp_a_months"
            )
        with c2:
            b_pick = st.multiselect(
                "Selection B (one or more months)",
                options=month_labels,
                default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                key="cmp_b_months"
            )
        with c3:
            by = st.selectbox("Compare by", ["Retailer", "Vendor"], key="cmp_by")

        if by == "Retailer":
            options = sorted(d["Retailer"].dropna().unique().tolist())
        else:
            options = sorted([v for v in d["Vendor"].dropna().unique().tolist() if str(v).strip()])

        sel = st.multiselect(f"Limit to {by}(s) (optional)", options=options, key="cmp_limit")

        a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
        b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

        if not a_periods or not b_periods:
            st.info("Pick at least one month in Selection A and Selection B.")
        else:
            da = d[d["MonthP"].isin(a_periods)]
            db = d[d["MonthP"].isin(b_periods)]

            if sel:
                da = da[da[by].isin(sel)]
                db = db[db[by].isin(sel)]

            ga = da.groupby(by, as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
            gb = db.groupby(by, as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))

            out = ga.merge(gb, on=by, how="outer").fillna(0.0)
            out["Units_Diff"] = out["Units_A"] - out["Units_B"]
            out["Sales_Diff"] = out["Sales_A"] - out["Sales_B"]
            out["Units_%"] = out["Units_Diff"] / out["Units_B"].replace(0, np.nan)
            out["Sales_%"] = out["Sales_Diff"] / out["Sales_B"].replace(0, np.nan)

            total = {
                by: "TOTAL",
                "Units_A": out["Units_A"].sum(),
                "Sales_A": out["Sales_A"].sum(),
                "Units_B": out["Units_B"].sum(),
                "Sales_B": out["Sales_B"].sum(),
            }
            total["Units_Diff"] = total["Units_A"] - total["Units_B"]
            total["Sales_Diff"] = total["Sales_A"] - total["Sales_B"]
            total["Units_%"] = total["Units_Diff"] / total["Units_B"] if total["Units_B"] else np.nan
            total["Sales_%"] = total["Sales_Diff"] / total["Sales_B"] if total["Sales_B"] else np.nan

            out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

            disp = out[[by,"Units_A","Sales_A","Units_B","Sales_B","Units_Diff","Units_%","Sales_Diff","Sales_%"]]
            sty = disp.style.format({
                "Units_A": fmt_int,
                "Units_B": fmt_int,
                "Units_Diff": fmt_int,
                "Units_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                "Sales_A": fmt_currency,
                "Sales_B": fmt_currency,
                "Sales_Diff": fmt_currency,
                "Sales_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
            }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])

            st.dataframe(sty, use_container_width=True, hide_index=True)
# -------------------------

# -------------------------
# Year Summary
# -------------------------
with tab_year_summary:
    st.subheader("Year Summary (YoY)")

    if df_all.empty:
        st.info("No sales data yet.")
    else:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)

        years = sorted(d["Year"].unique().tolist())
        c1, c2 = st.columns(2)
        with c1:
            base_year = st.selectbox("Base Year", options=years, index=0, key="ys_base")
        with c2:
            comp_year = st.selectbox("Comparison Year", options=years, index=(1 if len(years) > 1 else 0), key="ys_comp")

        basis = st.radio("Basis (tables + drivers)", options=["Sales", "Units"], index=0, horizontal=True, key="ys_basis")
        value_col = "Sales" if basis == "Sales" else "Units"

        a = d[d["Year"] == int(base_year)].copy()
        b = d[d["Year"] == int(comp_year)].copy()

        # KPIs
        def _sum(df_, col): 
            return float(df_[col].sum()) if not df_.empty else 0.0

        uA, uB = _sum(a, "Units"), _sum(b, "Units")
        sA, sB = _sum(a, "Sales"), _sum(b, "Sales")
        uD, sD = uB - uA, sB - sA
        uP = (uD / uA) if uA else np.nan
        sP = (sD / sA) if sA else np.nan

        st.markdown("### KPIs")
        k1, k2, k3, k4 = st.columns(4)
        k1.metric(f"Units ({base_year})", fmt_int(uA), delta=f"{fmt_int(uD)} ({uP*100:.1f}%)" if pd.notna(uP) else fmt_int(uD))
        k2.metric(f"Units ({comp_year})", fmt_int(uB))
        k3.metric(f"Sales ({base_year})", fmt_currency(sA), delta=f"{fmt_currency(sD)} ({sP*100:.1f}%)" if pd.notna(sP) else fmt_currency(sD))
        k4.metric(f"Sales ({comp_year})", fmt_currency(sB))

        # -------------------------
        # YoY driver breakdown (toggle)
        # -------------------------
        st.markdown("### YoY driver breakdown")

        sku_a = a.groupby("SKU", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
        sku_b = b.groupby("SKU", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))
        sku = sku_a.merge(sku_b, on="SKU", how="outer").fillna(0.0)

        sku["A_val"] = sku["Sales_A"] if value_col == "Sales" else sku["Units_A"]
        sku["B_val"] = sku["Sales_B"] if value_col == "Sales" else sku["Units_B"]
        sku["Delta"] = sku["B_val"] - sku["A_val"]

        sku["Bucket"] = "Same (flat)"
        sku.loc[(sku["A_val"] == 0) & (sku["B_val"] > 0), "Bucket"] = "New SKUs"
        sku.loc[(sku["A_val"] > 0) & (sku["B_val"] == 0), "Bucket"] = "Lost SKUs"
        sku.loc[(sku["A_val"] > 0) & (sku["B_val"] > 0) & (sku["Delta"] > 0), "Bucket"] = "Same SKUs – Growth"
        sku.loc[(sku["A_val"] > 0) & (sku["B_val"] > 0) & (sku["Delta"] < 0), "Bucket"] = "Same SKUs – Decline"

        total_delta = float(sku["Delta"].sum())

        def _b(name):
            return float(sku.loc[sku["Bucket"] == name, "Delta"].sum())

        b_new, b_lost = _b("New SKUs"), _b("Lost SKUs")
        b_grow, b_decl = _b("Same SKUs – Growth"), _b("Same SKUs – Decline")

        def _pct(x):
            return (x / total_delta) if total_delta else np.nan

        def _fmt(v):
            return fmt_currency(v) if value_col == "Sales" else fmt_int(v)

        cD1, cD2, cD3, cD4 = st.columns(4)
        cD1.metric("New SKUs", _fmt(b_new), delta=(f"{_pct(b_new)*100:.1f}%" if pd.notna(_pct(b_new)) else "—"))
        cD2.metric("Lost SKUs", _fmt(b_lost), delta=(f"{_pct(b_lost)*100:.1f}%" if pd.notna(_pct(b_lost)) else "—"))
        cD3.metric("Same SKUs – Growth", _fmt(b_grow), delta=(f"{_pct(b_grow)*100:.1f}%" if pd.notna(_pct(b_grow)) else "—"))
        cD4.metric("Same SKUs – Decline", _fmt(b_decl), delta=(f"{_pct(b_decl)*100:.1f}%" if pd.notna(_pct(b_decl)) else "—"))

        with st.expander("Top SKU drivers", expanded=False):
            tp = sku.sort_values("Delta", ascending=False).head(25).copy()
            tn = sku.sort_values("Delta", ascending=True).head(25).copy()

            tp2 = tp[["SKU","A_val","B_val","Delta","Bucket"]].rename(columns={"A_val": str(base_year), "B_val": str(comp_year)})
            tn2 = tn[["SKU","A_val","B_val","Delta","Bucket"]].rename(columns={"A_val": str(base_year), "B_val": str(comp_year)})

            st.markdown("**Top increases**")
            tp2_disp = tp2.copy()
            # Avoid Styler (pyarrow duplicate-column edge cases). Format numbers as strings.
            colA = str(base_year)
            colB = str(comp_year)
            if colA == colB:
                colA = f"{colA} (A)"
                colB = f"{colB} (B)"
                tp2_disp = tp2_disp.rename(columns={str(base_year): colA, str(comp_year): colB})
            num_cols = [c for c in [colA, colB, "Delta"] if c in tp2_disp.columns]
            for c in num_cols:
                tp2_disp[c] = tp2_disp[c].apply(_fmt)
            st.dataframe(tp2_disp, use_container_width=True, height=_table_height(tp2_disp, max_px=700), hide_index=True)

            st.markdown("**Top decreases**")
            tn2_disp = tn2.copy()
            colA = str(base_year)
            colB = str(comp_year)
            if colA == colB:
                colA = f"{colA} (A)"
                colB = f"{colB} (B)"
                tn2_disp = tn2_disp.rename(columns={str(base_year): colA, str(comp_year): colB})
            num_cols = [c for c in [colA, colB, "Delta"] if c in tn2_disp.columns]
            for c in num_cols:
                tn2_disp[c] = tn2_disp[c].apply(_fmt)
            st.dataframe(tn2_disp, use_container_width=True, height=_table_height(tn2_disp, max_px=700), hide_index=True)

        # -------------------------
        # Concentration risk (uses same toggle)
        # -------------------------
        st.markdown("### Concentration risk")

        def _top_share(df_year, group_col, topn):
            g = df_year.groupby(group_col, as_index=False).agg(val=(value_col, "sum"))
            total = float(g["val"].sum())
            if total <= 0:
                return 0.0
            return float(g.sort_values("val", ascending=False).head(topn)["val"].sum()) / total

        conc = pd.DataFrame([
            {
                "Year": int(base_year),
                "Top 1 Retailer %": _top_share(a, "Retailer", 1),
                "Top 3 Retailers %": _top_share(a, "Retailer", 3),
                "Top 5 Retailers %": _top_share(a, "Retailer", 5),
                "Top 1 Vendor %": _top_share(a, "Vendor", 1),
                "Top 3 Vendors %": _top_share(a, "Vendor", 3),
                "Top 5 Vendors %": _top_share(a, "Vendor", 5),
            },
            {
                "Year": int(comp_year),
                "Top 1 Retailer %": _top_share(b, "Retailer", 1),
                "Top 3 Retailers %": _top_share(b, "Retailer", 3),
                "Top 5 Retailers %": _top_share(b, "Retailer", 5),
                "Top 1 Vendor %": _top_share(b, "Vendor", 1),
                "Top 3 Vendors %": _top_share(b, "Vendor", 3),
                "Top 5 Vendors %": _top_share(b, "Vendor", 5),
            },
        ])

        conc_disp = conc.copy()
        st.dataframe(conc_disp.style.format({c: (lambda v: f"{v*100:.1f}%") for c in conc_disp.columns if c != "Year"}),
                     use_container_width=True, hide_index=True)

        # -------------------------
        # Retailer summary
        # -------------------------
        st.markdown("### Retailer summary")
        ra = a.groupby("Retailer", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
        rb = b.groupby("Retailer", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))
        r = ra.merge(rb, on="Retailer", how="outer").fillna(0.0)
        r["Units_Diff"] = r["Units_B"] - r["Units_A"]
        r["Sales_Diff"] = r["Sales_B"] - r["Sales_A"]
        r["Units_%"] = r["Units_Diff"] / r["Units_A"].replace(0, np.nan)
        r["Sales_%"] = r["Sales_Diff"] / r["Sales_A"].replace(0, np.nan)

        rsort = st.selectbox("Sort retailer table by", ["Sales_Diff","Units_Diff","Sales_B","Sales_A"], key="ys_r_sort")
        r = r.sort_values(rsort, ascending=False, kind="mergesort")
        r_disp = r[["Retailer","Units_A","Units_B","Units_Diff","Units_%","Sales_A","Sales_B","Sales_Diff","Sales_%"]]
        r_sty = r_disp.style.format({
            "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int,
            "Units_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
            "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency,
            "Sales_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
        }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])
        st.dataframe(r_sty, use_container_width=True, height=_table_height(r_disp, max_px=1100), hide_index=True)

        # -------------------------
        # Vendor summary
        # -------------------------
        st.markdown("### Vendor summary")
        va = a.groupby("Vendor", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
        vb = b.groupby("Vendor", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))
        v = va.merge(vb, on="Vendor", how="outer").fillna(0.0)
        v["Units_Diff"] = v["Units_B"] - v["Units_A"]
        v["Sales_Diff"] = v["Sales_B"] - v["Sales_A"]
        v["Units_%"] = v["Units_Diff"] / v["Units_A"].replace(0, np.nan)
        v["Sales_%"] = v["Sales_Diff"] / v["Sales_A"].replace(0, np.nan)

        vsort = st.selectbox("Sort vendor table by", ["Sales_Diff","Units_Diff","Sales_B","Sales_A"], key="ys_v_sort")
        v = v.sort_values(vsort, ascending=False, kind="mergesort")
        v_disp = v[["Vendor","Units_A","Units_B","Units_Diff","Units_%","Sales_A","Sales_B","Sales_Diff","Sales_%"]]
        v_sty = v_disp.style.format({
            "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int,
            "Units_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
            "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency,
            "Sales_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
        }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])
        st.dataframe(v_sty, use_container_width=True, height=_table_height(v_disp, max_px=1100), hide_index=True)


# -------------------------
# Data Inventory
# -------------------------
with tab_inventory:
    st.subheader("Data Inventory")

    if df_all.empty:
        st.info("No sales data yet.")
    else:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)

        st.markdown("### Rows by year")
        by_year = d.groupby("Year", as_index=False).agg(
            Rows=("SKU","size"),
            Units=("Units","sum"),
            Sales=("Sales","sum"),
            Retailers=("Retailer","nunique"),
            Vendors=("Vendor","nunique"),
            SKUs=("SKU","nunique"),
        ).sort_values("Year", ascending=False)
        st.dataframe(by_year.style.format({
            "Rows": fmt_int, "Units": fmt_int, "Sales": fmt_currency,
            "Retailers": fmt_int, "Vendors": fmt_int, "SKUs": fmt_int
        }), use_container_width=True, hide_index=True)

        st.markdown("### Rows by retailer (selected year)")
        years = sorted(d["Year"].unique().tolist())
        sel_y = st.selectbox("Year", options=years, index=len(years)-1, key="inv_year")
        dy = d[d["Year"] == int(sel_y)].copy()
        if "SourceFile" not in dy.columns:
            dy["SourceFile"] = ""
        by_ret = dy.groupby("Retailer", as_index=False).agg(
            Rows=("SKU","size"),
            Units=("Units","sum"),
            Sales=("Sales","sum"),
            SKUs=("SKU","nunique"),
            Sources=("SourceFile","nunique"),
        ).sort_values("Sales", ascending=False)
        st.dataframe(by_ret.style.format({
            "Rows": fmt_int, "Units": fmt_int, "Sales": fmt_currency, "SKUs": fmt_int, "Sources": fmt_int
        }), use_container_width=True, height=_table_height(by_ret, max_px=900), hide_index=True)

        st.markdown("### Rows by source file (selected year)")
        by_src = dy.groupby("SourceFile", as_index=False).agg(
            Rows=("SKU","size"),
            Units=("Units","sum"),
            Sales=("Sales","sum"),
            Retailers=("Retailer","nunique"),
            SKUs=("SKU","nunique"),
        ).sort_values("Sales", ascending=False)
        st.dataframe(by_src.style.format({
            "Rows": fmt_int, "Units": fmt_int, "Sales": fmt_currency, "Retailers": fmt_int, "SKUs": fmt_int
        }), use_container_width=True, height=_table_height(by_src, max_px=900), hide_index=True)




# -------------------------
# Insights & Alerts
# -------------------------
with tab_alerts:
    st.subheader("Insights & Alerts")

    if df_all.empty:
        st.info("No sales data yet.")
    else:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)
        years = sorted(d["Year"].unique().tolist())

        c1, c2 = st.columns(2)
        with c1:
            base_year = st.selectbox("Base Year", options=years, index=0, key="al_base")
        with c2:
            comp_year = st.selectbox("Comparison Year", options=years, index=(1 if len(years) > 1 else 0), key="al_comp")

        basis = st.radio("Basis", options=["Sales", "Units"], index=0, horizontal=True, key="al_basis")
        value_col = "Sales" if basis == "Sales" else "Units"

        a = d[d["Year"] == int(base_year)].copy()
        b = d[d["Year"] == int(comp_year)].copy()

        insights = []

        # Vendor deltas (worst 5)
        va = a.groupby("Vendor", as_index=False).agg(A=(value_col, "sum"))
        vb = b.groupby("Vendor", as_index=False).agg(B=(value_col, "sum"))
        v = va.merge(vb, on="Vendor", how="outer").fillna(0.0)
        v["Delta"] = v["B"] - v["A"]
        v = v.sort_values("Delta")

        def _fmt(vv):
            return fmt_currency(vv) if value_col == "Sales" else fmt_int(vv)

        for _, row in v.head(5).iterrows():
            if row["Delta"] < 0:
                insights.append(f"🔻 Vendor **{row['Vendor']}** down {_fmt(row['Delta'])} ({base_year} → {comp_year}).")

        # Retailer concentration warning (top 1 >= 40%)
        g = b.groupby("Retailer", as_index=False).agg(val=(value_col, "sum")).sort_values("val", ascending=False)
        total = float(g["val"].sum())
        if total > 0 and not g.empty:
            top1_share = float(g.iloc[0]["val"]) / total
            if top1_share >= 0.40:
                insights.append(f"⚠️ Concentration risk: **{g.iloc[0]['Retailer']}** is {top1_share*100:.1f}% of {comp_year} ({value_col}).")

        # Growth driven by few SKUs (top10 >= 60% of positive delta)
        sa = a.groupby("SKU", as_index=False).agg(A=(value_col, "sum"))
        sb = b.groupby("SKU", as_index=False).agg(B=(value_col, "sum"))
        sku = sa.merge(sb, on="SKU", how="outer").fillna(0.0)
        sku["Delta"] = sku["B"] - sku["A"]

        pos = sku[sku["Delta"] > 0].sort_values("Delta", ascending=False)
        if not pos.empty:
            top10 = float(pos.head(10)["Delta"].sum())
            total_pos = float(pos["Delta"].sum())
            share = (top10 / total_pos) if total_pos else 0.0
            if share >= 0.60:
                insights.append(f"📈 Growth concentration: top 10 SKUs drive {share*100:.1f}% of positive YoY change ({value_col}).")

        # Lost SKUs count
        lost = int(((sku["A"] > 0) & (sku["B"] == 0)).sum())
        if lost:
            insights.append(f"🧯 Lost SKUs: **{lost}** SKUs sold in {base_year} but not in {comp_year}.")

        # Year locks notice
        locked = sorted(list(load_year_locks()))
        if locked:
            insights.append(f"🔒 Locked years: {', '.join(str(y) for y in locked)} (bulk ingest blocked).")

        if not insights:
            st.success("No major alerts detected with the current settings.")
        else:
            st.markdown("### Highlights")
            for s in insights:
                st.markdown(f"- {s}")

        with st.expander("Details (tables)", expanded=False):
            st.markdown("**Worst vendors**")
            st.dataframe(v.head(15).style.format({"A": _fmt, "B": _fmt, "Delta": _fmt})
                         .applymap(lambda v: f"color: {_color(v)};", subset=["Delta"]),
                         use_container_width=True, hide_index=True)

            st.markdown("**Top SKU movers**")
            movers = sku.sort_values("Delta", ascending=False).head(15).copy()
            st.dataframe(movers.style.format({"A": _fmt, "B": _fmt, "Delta": _fmt})
                         .applymap(lambda v: f"color: {_color(v)};", subset=["Delta"]),
                         use_container_width=True, hide_index=True)

# Run-Rate Forecast
# -------------------------
with tab_runrate:
    st.subheader("Run-Rate Forecast")

    if df.empty:
        st.info("No sales data yet.")
    else:
        window = st.selectbox("Forecast window (weeks)", options=[4, 8, 12], index=0, key="rr_window")
        lookback = st.selectbox("Lookback for avg", options=[4, 8, 12], index=1, key="rr_lookback")
        level = st.selectbox("Level", options=["SKU", "Vendor", "Retailer"], index=0, key="rr_level")

        d = add_week_col(df)
        weeks = last_n_weeks(d, lookback)
        d = d[d["Week"].isin(weeks)].copy()

        if level == "SKU":
            grp = ["Retailer","Vendor","SKU"]
        elif level == "Vendor":
            grp = ["Vendor"]
        else:
            grp = ["Retailer"]

        base = d.groupby(grp + ["Week"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
        units_piv = base.pivot_table(index=grp, columns="Week", values="Units", aggfunc="sum", fill_value=0.0)
        sales_piv = base.pivot_table(index=grp, columns="Week", values="Sales", aggfunc="sum", fill_value=0.0)

        avg_units = nonzero_mean_rowwise(units_piv).fillna(0.0)
        avg_sales = nonzero_mean_rowwise(sales_piv).fillna(0.0)

        out = avg_units.reset_index().rename(columns={0:"AvgWeeklyUnits"})
        out["AvgWeeklySales"] = avg_sales.values
        out["ProjectedUnits"] = out["AvgWeeklyUnits"] * window
        out["ProjectedSales"] = out["AvgWeeklySales"] * window
        out = out.sort_values("ProjectedSales", ascending=False)

        disp = out.copy()
        disp["AvgWeeklyUnits"] = disp["AvgWeeklyUnits"].round(2)
        disp["ProjectedUnits"] = disp["ProjectedUnits"].round(0).astype(int)

        sty = disp.style.format({
            "AvgWeeklyUnits": lambda v: fmt_2(v),
            "AvgWeeklySales": lambda v: fmt_currency(v),
            "ProjectedUnits": lambda v: fmt_int(v),
            "ProjectedSales": lambda v: fmt_currency(v),
        })
        st.dataframe(sty, use_container_width=True, height=_table_height(disp, max_px=1200), hide_index=True)

# -------------------------
# Seasonality Heatmap
# -------------------------

with tab_exec:
    st.subheader("Executive Summary")

    scope = st.selectbox("Scope", options=["All", "Retailer", "Vendor"], index=0, key="ex_scope")
    if scope == "Retailer":
        opts = sorted(vmap["Retailer"].dropna().unique().tolist())
        pick = st.selectbox("Retailer", options=opts, index=0, key="ex_pick_r")
        d = df[df["Retailer"] == pick].copy()
        title = f"Executive Summary - {pick}"
    elif scope == "Vendor":
        opts = sorted([v for v in vmap["Vendor"].dropna().unique().tolist() if str(v).strip() != ""])
        pick = st.selectbox("Vendor", options=opts, index=0, key="ex_pick_v")
        d = df[df["Vendor"] == pick].copy()
        title = f"Executive Summary - {pick}"
    else:
        d = df.copy()
        title = "Executive Summary - All Retailers"

    d = add_week_col(d)

    tf_label = st.selectbox(
        "Timeframe",
        options=["2 weeks","4 weeks","6 weeks","8 weeks","3 months","6 months","12 months","YTD (all loaded)"],
        index=1,
        key="ex_tf2"
    )

    # Apply timeframe
    if "weeks" in tf_label:
        n = int(tf_label.split()[0])
        use = last_n_weeks(d, n)
        d = d[d["Week"].isin(use)].copy()
    elif "months" in tf_label:
        n = int(tf_label.split()[0])
        d["MonthP"] = pd.to_datetime(d["StartDate"], errors="coerce").dt.to_period("M")
        months = sorted(d["MonthP"].dropna().unique().tolist())
        usem = months[-n:] if len(months) >= n else months
        d = d[d["MonthP"].isin(usem)].copy()
    else:
        pass

    m = wow_mom_metrics(d)
    cols = st.columns(6)
    cols[0].metric("Units", fmt_int(m["total_units"]))
    cols[1].metric("Sales", fmt_currency(m["total_sales"]))
    cols[2].markdown(f"<div style='color:{_color(m['wow_units'])}; font-weight:600;'>WoW Units: {fmt_int(m['wow_units']) if m['wow_units'] is not None else '—'}</div>", unsafe_allow_html=True)
    cols[3].markdown(f"<div style='color:{_color(m['wow_sales'])}; font-weight:600;'>WoW Sales: {fmt_currency(m['wow_sales']) if m['wow_sales'] is not None else '—'}</div>", unsafe_allow_html=True)
    cols[4].markdown(f"<div style='color:{_color(m['mom_units'])}; font-weight:600;'>MoM Units: {fmt_int(m['mom_units']) if m['mom_units'] is not None else '—'}</div>", unsafe_allow_html=True)
    cols[5].markdown(f"<div style='color:{_color(m['mom_sales'])}; font-weight:600;'>MoM Sales: {fmt_currency(m['mom_sales']) if m['mom_sales'] is not None else '—'}</div>", unsafe_allow_html=True)

    st.divider()

    # Monthly totals table
    d2 = d.copy()
    d2["StartDate"] = pd.to_datetime(d2["StartDate"], errors="coerce")
    d2["MonthP"] = d2["StartDate"].dt.to_period("M")
    mon = d2.groupby("MonthP", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("MonthP")
    if not mon.empty:
        mon["Month"] = mon["MonthP"].map(month_label)
        mon = mon[["Month","Units","Sales"]]
        mon["Units"] = mon["Units"].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
        st.markdown("### Monthly totals")
        st.dataframe(mon.style.format({"Units": lambda v: fmt_int(v), "Sales": lambda v: fmt_currency(v)}),
                     use_container_width=True, height=_table_height(mon, max_px=800), hide_index=True)

    # Mix table depending on scope
    if scope == "Retailer":
        mix = d.groupby("Vendor", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
        mix = mix[(mix["Units"].fillna(0) > 0) | (mix["Sales"].fillna(0) > 0)]
        total_u = float(mix["Units"].sum()) if not mix.empty else 0.0
        total_s = float(mix["Sales"].sum()) if not mix.empty else 0.0
        mix["% Units"] = mix["Units"].apply(lambda v: (v/total_u) if total_u else 0.0)
        mix["% Sales"] = mix["Sales"].apply(lambda v: (v/total_s) if total_s else 0.0)
        mix = mix.sort_values("% Sales", ascending=False)
        mix["Units"] = mix["Units"].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
        st.markdown("### Vendor mix for this retailer")
        st.dataframe(mix.style.format({
            "Units": lambda v: fmt_int(v),
            "Sales": lambda v: fmt_currency(v),
            "% Units": lambda v: f"{v*100:.1f}%",
            "% Sales": lambda v: f"{v*100:.1f}%"
        }), use_container_width=True, height=_table_height(mix, max_px=900), hide_index=True)

    elif scope == "Vendor":
        mix = d.groupby("Retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
        mix = mix[(mix["Units"].fillna(0) > 0) | (mix["Sales"].fillna(0) > 0)]
        total_u = float(mix["Units"].sum()) if not mix.empty else 0.0
        total_s = float(mix["Sales"].sum()) if not mix.empty else 0.0
        mix["% Units"] = mix["Units"].apply(lambda v: (v/total_u) if total_u else 0.0)
        mix["% Sales"] = mix["Sales"].apply(lambda v: (v/total_s) if total_s else 0.0)
        mix = mix.sort_values("% Sales", ascending=False)
        mix["Units"] = mix["Units"].map(lambda v: int(round(float(v))) if pd.notna(v) else 0)
        st.markdown("### Retailer mix for this vendor")
        st.dataframe(mix.style.format({
            "Units": lambda v: fmt_int(v),
            "Sales": lambda v: fmt_currency(v),
            "% Units": lambda v: f"{v*100:.1f}%",
            "% Sales": lambda v: f"{v*100:.1f}%"
        }), use_container_width=True, height=_table_height(mix, max_px=900), hide_index=True)

    st.divider()

    sku_agg = d.groupby(["SKU","Retailer","Vendor"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
    sku_agg["Vendor"] = sku_agg["Vendor"].fillna("Unmapped")
    sold = sku_agg[(sku_agg["Units"].fillna(0) > 0) & (sku_agg["Sales"].fillna(0) > 0)].copy()

    top_units = sold.sort_values("Units", ascending=False).head(10)[["SKU","Retailer","Vendor","Units"]].copy()
    top_sales = sold.sort_values("Sales", ascending=False).head(10)[["SKU","Retailer","Vendor","Sales"]].copy()
    bot_units = sold.sort_values("Units", ascending=True).head(10)[["SKU","Retailer","Vendor","Units"]].copy()
    bot_sales = sold.sort_values("Sales", ascending=True).head(10)[["SKU","Retailer","Vendor","Sales"]].copy()

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("### Top 10 SKUs by Units")
        tu = top_units.copy()
        tu["Units"] = tu["Units"].map(lambda v: int(round(float(v))))
        st.dataframe(tu.style.format({"Units": lambda v: fmt_int(v)}),
                     use_container_width=True, height=_table_height(tu, max_px=700), hide_index=True)

        st.markdown("### Bottom 10 SKUs by Units")
        bu = bot_units.copy()
        bu["Units"] = bu["Units"].map(lambda v: int(round(float(v))))
        st.dataframe(bu.style.format({"Units": lambda v: fmt_int(v)}),
                     use_container_width=True, height=_table_height(bu, max_px=700), hide_index=True)

    with c2:
        st.markdown("### Top 10 SKUs by Sales")
        st.dataframe(top_sales.style.format({"Sales": lambda v: fmt_currency(v)}),
                     use_container_width=True, height=_table_height(top_sales, max_px=700), hide_index=True)

        st.markdown("### Bottom 10 SKUs by Sales")
        st.dataframe(bot_sales.style.format({"Sales": lambda v: fmt_currency(v)}),
                     use_container_width=True, height=_table_height(bot_sales, max_px=700), hide_index=True)

    st.divider()

    summary_rows = [
        {"Metric":"Total Units", "Value": m["total_units"]},
        {"Metric":"Total Sales", "Value": m["total_sales"]},
        {"Metric":"WoW Units", "Value": m["wow_units"]},
        {"Metric":"WoW Sales", "Value": m["wow_sales"]},
        {"Metric":"MoM Units", "Value": m["mom_units"]},
        {"Metric":"MoM Sales", "Value": m["mom_sales"]},
    ]
    summary_df = pd.DataFrame(summary_rows)
    st.download_button("Download KPIs (CSV)", data=summary_df.to_csv(index=False).encode("utf-8"),
                       file_name="executive_kpis.csv", mime="text/csv")

    sections = [
        ("KPIs", [
            f"Total Units: {fmt_int(m['total_units'])}",
            f"Total Sales: {fmt_currency(m['total_sales'])}",
            f"WoW Units: {fmt_int(m['wow_units']) if m['wow_units'] is not None else '—'}",
            f"WoW Sales: {fmt_currency(m['wow_sales']) if m['wow_sales'] is not None else '—'}",
            f"MoM Units: {fmt_int(m['mom_units']) if m['mom_units'] is not None else '—'}",
            f"MoM Sales: {fmt_currency(m['mom_sales']) if m['mom_sales'] is not None else '—'}",
        ]),
        ("Top 10 SKUs by Units", [f"{r.SKU} | {r.Retailer} | {r.Vendor} | {int(round(r.Units))}" for r in top_units.itertuples(index=False)]),
        ("Top 10 SKUs by Sales", [f"{r.SKU} | {r.Retailer} | {r.Vendor} | {fmt_currency(r.Sales)}" for r in top_sales.itertuples(index=False)]),
    ]
    pdf_bytes = to_pdf_bytes(title, sections)
    if pdf_bytes:
        st.download_button("Download Executive Summary (PDF)", data=pdf_bytes,
                           file_name="executive_summary.pdf", mime="application/pdf")
    else:
        st.info("PDF export requires the reportlab package.")

# Edit Vendor Map
with tab_edit_map:
    st.subheader("Edit Vendor Map")
    st.caption("Edit Vendor and Price. Click Save to update the default vendor map file used by the app.")
    vmap_disp = vmap[["Retailer","SKU","Vendor","Price","MapOrder"]].copy().sort_values(["Retailer","MapOrder"])
    show = vmap_disp.drop(columns=["MapOrder"]).copy()

    if edit_mode:
        edited = st.data_editor(show, use_container_width=True, hide_index=True, num_rows="dynamic")
        if st.button("Save Vendor Map"):
            updated = edited.copy()
            updated["Retailer"] = updated["Retailer"].map(_normalize_retailer)
            updated["SKU"] = updated["SKU"].map(_normalize_sku)
            updated["Vendor"] = updated["Vendor"].astype(str).str.strip()
            updated["Price"] = pd.to_numeric(updated["Price"], errors="coerce")

            # MapOrder based on current row order per retailer
            updated["MapOrder"] = 0
            for r, grp in updated.groupby("Retailer", sort=False):
                for j, ix in enumerate(grp.index.tolist()):
                    updated.loc[ix, "MapOrder"] = j

            updated.to_excel(DEFAULT_VENDOR_MAP, index=False)
            st.success("Saved vendor map. Reloading…")
            st.rerun()
    else:
        st.info("Turn on Edit Mode in the sidebar to edit.")
        st.dataframe(show, use_container_width=True, height=_table_height(show, max_px=1400), hide_index=True)

# Backup / Restore
with tab_backup:
    st.subheader("Backup / Restore")

    st.markdown("### Backup files")
    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("#### Sales database")
        if DEFAULT_SALES_STORE.exists():
            st.download_button("Download sales_store.csv", data=DEFAULT_SALES_STORE.read_bytes(), file_name="sales_store.csv", mime="text/csv")
        else:
            st.info("No sales_store.csv yet.")

        up = st.file_uploader("Restore sales_store.csv", type=["csv"], key="restore_sales_csv")
        if st.button("Restore sales_store.csv", disabled=up is None, key="btn_restore_sales"):
            DEFAULT_SALES_STORE.write_bytes(up.getbuffer())
            st.success("Restored sales_store.csv. Reloading…")
            st.rerun()

    with c2:
        st.markdown("#### Vendor map")
        if DEFAULT_VENDOR_MAP.exists():
            st.download_button("Download vendor_map.xlsx", data=DEFAULT_VENDOR_MAP.read_bytes(), file_name="vendor_map.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("No vendor_map.xlsx yet.")

        up2 = st.file_uploader("Restore vendor_map.xlsx", type=["xlsx"], key="restore_vm_xlsx")
        if st.button("Restore vendor_map.xlsx", disabled=up2 is None, key="btn_restore_vm"):
            DEFAULT_VENDOR_MAP.write_bytes(up2.getbuffer())
            st.success("Restored vendor_map.xlsx. Reloading…")
            st.rerun()

    with c3:
        st.markdown("#### Price history")
        if DEFAULT_PRICE_HISTORY.exists():
            st.download_button("Download price_history.csv", data=DEFAULT_PRICE_HISTORY.read_bytes(), file_name="price_history.csv", mime="text/csv")
        else:
            st.info("No price_history.csv yet.")

        up3 = st.file_uploader("Restore price_history.csv", type=["csv"], key="restore_ph_csv")
        if st.button("Restore price_history.csv", disabled=up3 is None, key="btn_restore_ph"):
            DEFAULT_PRICE_HISTORY.write_bytes(up3.getbuffer())
            st.success("Restored price_history.csv. Reloading…")
            st.rerun()

    st.markdown("#### Year locks")
    if DEFAULT_YEAR_LOCKS.exists():
        st.download_button("Download year_locks.json", data=DEFAULT_YEAR_LOCKS.read_bytes(), file_name="year_locks.json", mime="application/json")
    else:
        st.info("No year locks saved yet.")

    up4 = st.file_uploader("Restore year_locks.json", type=["json"], key="restore_year_locks")
    if st.button("Restore year_locks.json", disabled=up4 is None, key="btn_restore_year_locks"):
        DEFAULT_YEAR_LOCKS.write_bytes(up4.getbuffer())
        st.success("Restored year locks. Reloading…")
        st.rerun()

    st.divider()

    st.markdown("### Price changes (effective date)")
    st.caption("Upload a sheet with SKU + Price + StartDate. Optional Retailer column. Prices apply from StartDate forward and never change earlier weeks.")

    tmpl = pd.DataFrame([
        {"Retailer":"*", "SKU":"ABC123", "Price": 19.99, "StartDate":"2026-02-01"},
        {"Retailer":"home depot", "SKU":"XYZ999", "Price": 24.99, "StartDate":"2026-03-15"},
    ])
    st.download_button("Download template CSV", data=tmpl.to_csv(index=False).encode("utf-8"),
                       file_name="price_history_template.csv", mime="text/csv")

    ph_up = st.file_uploader("Upload price history (CSV or Excel)", type=["csv","xlsx"], key="ph_upload")
    if ph_up is not None:
        try:
            if ph_up.name.lower().endswith(".csv"):
                ph_new = pd.read_csv(ph_up)
            else:
                ph_new = pd.read_excel(ph_up)

            st.markdown("#### Preview upload")
            st.dataframe(ph_new.head(50), use_container_width=True, hide_index=True)

            # Normalize + ignore blanks safely
            cur_ph = load_price_history()
            incoming, ignored = _prepare_price_history_upload(ph_new)
            diff = _price_history_diff(cur_ph, incoming)

            st.divider()
            st.markdown("#### What will change")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Rows uploaded", int(len(ph_new)))
            c2.metric("Rows ignored (blank/invalid)", int(len(ignored)))
            c3.metric("Inserts", int((diff["Action"] == "insert").sum()) if not diff.empty else 0)
            c4.metric("Updates", int((diff["Action"] == "update").sum()) if not diff.empty else 0)

            show_diff = diff.copy()
            if not show_diff.empty:
                show_diff["StartDate"] = pd.to_datetime(show_diff["StartDate"], errors="coerce").dt.date
                sty = show_diff.style.format({
                    "OldPrice": lambda v: fmt_currency(v) if pd.notna(v) else "—",
                    "Price": lambda v: fmt_currency(v),
                    "PriceDiff": lambda v: fmt_currency(v) if pd.notna(v) else "—",
                }).applymap(lambda v: "font-weight:700;" if str(v) in ["insert","update"] else "", subset=["Action"])
                st.dataframe(sty, use_container_width=True, height=_table_height(show_diff, max_px=900), hide_index=True)

                st.download_button("Download change preview (CSV)", data=show_diff.to_csv(index=False).encode("utf-8"),
                                   file_name="price_history_changes_preview.csv", mime="text/csv")
            else:
                st.info("No valid rows found in this upload (all prices were blank/invalid).")

            if not ignored.empty:
                st.markdown("#### Ignored rows")
                ign = ignored.copy()
                ign["StartDate"] = pd.to_datetime(ign["StartDate"], errors="coerce").dt.date
                st.dataframe(ign.head(200), use_container_width=True, height=_table_height(ign, max_px=600), hide_index=True)
                st.download_button("Download ignored rows (CSV)", data=ign.to_csv(index=False).encode("utf-8"),
                                   file_name="price_history_ignored_rows.csv", mime="text/csv")

            if st.button("Apply price changes", key="btn_apply_prices"):
                ins, upd, noop = upsert_price_history(ph_new)
                st.success(f"Price history updated. Inserts: {ins}, Updates: {upd}, Unchanged: {noop}. Reloading…")
                st.rerun()
        except Exception as e:
            st.error(f"Could not read this file: {e}")

    if DEFAULT_PRICE_HISTORY.exists():
        if st.button("Clear ALL price history", key="btn_clear_ph"):
            DEFAULT_PRICE_HISTORY.unlink(missing_ok=True)
            st.success("Cleared. Reloading…")
            st.rerun()

    st.divider()

    st.markdown("### Export enriched sales")
    if not df.empty:
        ex = df.copy()
        ex["StartDate"] = pd.to_datetime(ex["StartDate"], errors="coerce").dt.strftime("%Y-%m-%d")
        ex["EndDate"] = pd.to_datetime(ex["EndDate"], errors="coerce").dt.strftime("%Y-%m-%d")
        st.download_button("Download enriched_sales.csv", data=ex.to_csv(index=False).encode("utf-8"),
                           file_name="enriched_sales.csv", mime="text/csv")
    else:
        st.info("No sales yet.")



# -------------------------
# Bulk Data Upload
# -------------------------
with tab_bulk_upload:
    st.subheader("Bulk Data Upload (Multi-week / Multi-month)")

    st.markdown(
        """
        Use this when you get a **wide** retailer file (not week-by-week uploads).

        Expected format:
        - One sheet per retailer (or retailer name in cell **A1**)
        - Column **A** = SKU (starting row 2)
        - Row **1** from column **B** onward = week ranges (example: `1-1 / 1-3`)
        - Cells = Units sold for that SKU in that week
        - Sales uses your **current pricing** (Vendor Map / Price History). `UnitPrice` is left blank.
        """
    )

    locked_years = load_year_locks()
    years_opt = list(range(this_year - 6, this_year + 2))

    st.markdown("### Year locks")
    cL1, cL2 = st.columns([2, 1])
    with cL1:
        lock_pick = st.multiselect("Locked years (prevent edits)", options=years_opt, default=sorted(list(locked_years)), key="lock_pick")
    with cL2:
        if st.button("Save locks", key="btn_save_locks"):
            save_year_locks(set(int(y) for y in lock_pick))
            st.success("Saved year locks.")
            st.rerun()

    st.divider()

    bulk_upload = st.file_uploader(
        "Upload bulk data workbook (.xlsx)",
        type=["xlsx"],
        key="bulk_up_tab"
    )

    data_year = st.selectbox(
        "Data Year (for header parsing)",
        options=years_opt,
        index=years_opt.index(int(view_year)) if int(view_year) in years_opt else years_opt.index(this_year),
        key="bulk_data_year"
    )

    mode = st.radio(
        "Ingest mode",
        options=["Append (add rows)", "Overwrite year + retailer(s) (replace)"],
        index=0,
        horizontal=True,
        key="bulk_mode"
    )

    is_locked = int(data_year) in load_year_locks()
    if is_locked:
        st.error(f"Year {int(data_year)} is locked. Unlock it above to ingest data for this year.")

    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button("Ingest Bulk Workbook", disabled=(bulk_upload is None) or is_locked, key="btn_ingest_bulk"):
            new_rows = read_yow_workbook(bulk_upload, year=int(data_year))

            if mode.startswith("Overwrite"):
                retailers = set(new_rows["Retailer"].dropna().unique().tolist()) if not new_rows.empty else set()
                overwrite_sales_rows(int(data_year), retailers)

            append_sales_to_store(new_rows)
            st.success("Bulk workbook ingested successfully.")
            st.rerun()

    with c2:
        st.caption("Append = adds rows. Overwrite = deletes existing rows for that year + retailer(s) found in the upload, then re-adds.")

